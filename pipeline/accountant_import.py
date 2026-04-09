"""
Import accountant's monthly financial package (P&L, Balance Sheet, Statement of Cash Flows).

Expected Excel format:
  - Locations tab: maps location names to 2-letter codes
  - PL tab: Consolidated Profit & Loss (monthly columns)
  - BS tab: Consolidated Balance Sheet (monthly columns)
  - SCF tab: Statement of Cash Flows (monthly columns)
  - Studio tabs (BK, CC, DN, etc.): Per-studio P&L matching consolidated row structure

Usage:
    from pipeline.accountant_import import import_financials
    result = import_financials("path/to/file.xlsx")
"""

import json
import re
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


OUTPUT_DIR = Path(__file__).parent.parent / "data" / "financials"


def _parse_months_from_header(ws):
    """Extract month columns from header row (row 5)."""
    months = {}
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=5, column=col).value
        if val and isinstance(val, str):
            # Skip 'Total' column
            if val.strip().lower() == "total":
                continue
            months[col] = val.strip()
    return months


def _detect_period(ws):
    """Detect the reporting period from the date header (row 3)."""
    header = ws.cell(row=3, column=1).value or ""
    # e.g. "January, 2025-February, 2026" or "As of February 28, 2026"
    return header.strip()


def _last_month_from_header(months):
    """Get the last month string from the column headers."""
    if not months:
        return None
    last_col = max(months.keys())
    return months[last_col]


def _parse_sheet_to_df(ws, months):
    """Parse a sheet into a DataFrame with account labels as index and months as columns."""
    rows = []
    for row_idx in range(6, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label is None:
            continue
        label = str(label).strip()
        values = {}
        for col, month_name in months.items():
            cell_val = ws.cell(row=row_idx, column=col).value
            values[month_name] = cell_val if cell_val is not None else 0
        rows.append({"Account": label, **values})

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.set_index("Account")
    # Convert all value columns to numeric
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    return df


def _parse_locations(ws):
    """Parse the Locations tab into a dict: code -> name."""
    locations = {}
    for row in range(2, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        if name and code:
            locations[str(code).strip()] = str(name).strip()
    return locations


def import_financials(filepath, save=True):
    """
    Import the accountant's financial package from Excel.

    Args:
        filepath: Path to the accountant's Excel file.
        save: If True, save parsed data to data/financials/.

    Returns:
        dict with keys: pl, bs, scf, studios, locations, metadata
    """
    filepath = Path(filepath)
    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames

    # Parse locations
    locations = {}
    if "Locations" in sheet_names:
        locations = _parse_locations(wb["Locations"])

    # Parse consolidated sheets
    result = {"locations": locations, "studios": {}}

    for sheet_key, sheet_name in [("pl", "PL"), ("bs", "BS"), ("scf", "SCF")]:
        if sheet_name in sheet_names:
            ws = wb[sheet_name]
            months = _parse_months_from_header(ws)
            result[sheet_key] = _parse_sheet_to_df(ws, months)
        else:
            result[sheet_key] = pd.DataFrame()

    # Detect last actuals month from PL header
    if "PL" in sheet_names:
        pl_months = _parse_months_from_header(wb["PL"])
        last_month = _last_month_from_header(pl_months)
        period = _detect_period(wb["PL"])
    else:
        last_month = None
        period = ""

    # Parse studio tabs — any sheet whose name matches a location code
    known_codes = set(locations.keys())
    # Also include tabs that look like studio codes (2-3 letter uppercase)
    for sn in sheet_names:
        if sn in ("Locations", "PL", "BS", "SCF"):
            continue
        code = sn.strip()
        if code in known_codes or (len(code) <= 3 and code.isalpha() and code.isupper()):
            ws = wb[code]
            months = _parse_months_from_header(ws)
            studio_name = locations.get(code, code)
            df = _parse_sheet_to_df(ws, months)
            if not df.empty:
                result["studios"][code] = {
                    "name": studio_name,
                    "data": df,
                }

    # Metadata
    result["metadata"] = {
        "source_file": filepath.name,
        "imported_at": datetime.now().isoformat(),
        "period": period,
        "last_actuals_month": last_month,
        "studio_count": len(result["studios"]),
        "location_count": len(locations),
    }

    if save:
        _save_results(result)

    return result


def _save_results(result):
    """Save parsed data to data/financials/ as CSV + metadata JSON."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    last_month = result["metadata"].get("last_actuals_month", "unknown")
    # Clean month string for filenames: "February 2026" -> "Feb2026"
    month_tag = _month_tag(last_month)

    # Save consolidated sheets
    for key in ("pl", "bs", "scf"):
        df = result[key]
        if not df.empty:
            path = OUTPUT_DIR / f"{key}_{month_tag}.csv"
            df.to_csv(path)
            print(f"  Saved {key.upper()}: {path}")

    # Save studio P&Ls
    studio_dir = OUTPUT_DIR / f"studios_{month_tag}"
    studio_dir.mkdir(exist_ok=True)
    for code, studio in result["studios"].items():
        path = studio_dir / f"{code}_{month_tag}.csv"
        studio["data"].to_csv(path)

    if result["studios"]:
        print(f"  Saved {len(result['studios'])} studio P&Ls: {studio_dir}/")

    # Save metadata
    meta_path = OUTPUT_DIR / f"metadata_{month_tag}.json"
    with open(meta_path, "w") as f:
        json.dump(result["metadata"], f, indent=2)
    print(f"  Saved metadata: {meta_path}")

    # Save a "latest" symlink/pointer
    latest_path = OUTPUT_DIR / "latest.json"
    latest = {
        "month_tag": month_tag,
        "last_actuals_month": last_month,
        "imported_at": result["metadata"]["imported_at"],
        "source_file": result["metadata"]["source_file"],
    }
    with open(latest_path, "w") as f:
        json.dump(latest, f, indent=2)


def _month_tag(month_str):
    """Convert 'February 2026' -> 'Feb2026'."""
    if not month_str:
        return "unknown"
    parts = month_str.strip().split()
    if len(parts) == 2:
        return parts[0][:3] + parts[1]
    return month_str.replace(" ", "")


def load_latest():
    """Load the most recently imported financials from saved CSVs."""
    latest_path = OUTPUT_DIR / "latest.json"
    if not latest_path.exists():
        raise FileNotFoundError("No imported financials found. Run import-financials first.")

    with open(latest_path) as f:
        latest = json.load(f)

    month_tag = latest["month_tag"]
    result = {"metadata": latest, "studios": {}}

    for key in ("pl", "bs", "scf"):
        path = OUTPUT_DIR / f"{key}_{month_tag}.csv"
        if path.exists():
            result[key] = pd.read_csv(path, index_col=0)
        else:
            result[key] = pd.DataFrame()

    studio_dir = OUTPUT_DIR / f"studios_{month_tag}"
    if studio_dir.exists():
        for csv_file in sorted(studio_dir.glob("*.csv")):
            code = csv_file.stem.split("_")[0]
            result["studios"][code] = {
                "name": code,
                "data": pd.read_csv(csv_file, index_col=0),
            }

    return result


def print_summary(result):
    """Print a human-readable summary of imported financials."""
    meta = result["metadata"]
    print(f"\n{'='*60}")
    print(f"ACCOUNTANT FINANCIAL IMPORT")
    print(f"{'='*60}")
    print(f"  Source:       {meta['source_file']}")
    print(f"  Period:       {meta.get('period', 'N/A')}")
    print(f"  Last month:   {meta.get('last_actuals_month', 'N/A')}")
    print(f"  Imported:     {meta['imported_at'][:19]}")
    print()

    # P&L summary
    if not result["pl"].empty:
        pl = result["pl"]
        last_col = pl.columns[-1]
        total_rev = pl.loc["Total for Income", last_col] if "Total for Income" in pl.index else 0
        gross_profit = pl.loc["Gross Profit", last_col] if "Gross Profit" in pl.index else 0
        net_income = pl.loc["Net Income", last_col] if "Net Income" in pl.index else 0
        print(f"  P&L ({last_col}):")
        print(f"    Revenue:      ${total_rev:>12,.2f}")
        print(f"    Gross Profit: ${gross_profit:>12,.2f}")
        print(f"    Net Income:   ${net_income:>12,.2f}")
        print()

    # BS summary
    if not result["bs"].empty:
        bs = result["bs"]
        last_col = bs.columns[-1]
        total_assets = bs.loc["Total for Assets", last_col] if "Total for Assets" in bs.index else 0
        total_liab = bs.loc["Total for Liabilities", last_col] if "Total for Liabilities" in bs.index else 0
        cash = bs.loc["Total for Bank Accounts", last_col] if "Total for Bank Accounts" in bs.index else 0
        print(f"  Balance Sheet ({last_col}):")
        print(f"    Cash:             ${cash:>12,.2f}")
        print(f"    Total Assets:     ${total_assets:>12,.2f}")
        print(f"    Total Liabilities:${total_liab:>12,.2f}")
        print()

    # SCF summary
    if not result["scf"].empty:
        scf = result["scf"]
        last_col = scf.columns[-1]
        net_cash = scf.loc["NET CASH INCREASE FOR PERIOD", last_col] if "NET CASH INCREASE FOR PERIOD" in scf.index else 0
        print(f"  Cash Flow ({last_col}):")
        print(f"    Net Change:   ${net_cash:>12,.2f}")
        print()

    # Studios
    if result["studios"]:
        print(f"  Studios ({len(result['studios'])} locations):")
        for code in sorted(result["studios"].keys()):
            studio = result["studios"][code]
            df = studio["data"]
            last_col = df.columns[-1] if not df.empty else "N/A"
            rev = 0
            ni = 0
            if not df.empty:
                rev = df.loc["Total Income", last_col] if "Total Income" in df.index else 0
                ni = df.loc["Net Income", last_col] if "Net Income" in df.index else 0
            name = studio["name"]
            if rev == 0 and ni == 0:
                print(f"    {code:3s} {name:20s}  (placeholder)")
            else:
                print(f"    {code:3s} {name:20s}  Rev: ${rev:>10,.0f}  NI: ${ni:>10,.0f}")

    print(f"\n{'='*60}")
