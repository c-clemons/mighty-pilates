"""
Frozen monthly GL — bit-exact reproducibility for closed months.

Two tables:
  FROZEN_MONTHLY_GL: (MONTH_YM, STUDIO_NAME, GL_CODE) → AMOUNT
  SERVICE_TYPE_BUCKETS_FROZEN: (MONTH_YM, SERVICE_TYPE) → BUCKET

Sign convention stored in FROZEN_MONTHLY_GL is the **GL convention**:
  - Revenue accounts (401*, 403*, 404*): POSITIVE
  - Contra accounts (406000 Refunds, 407000 Discounts): NEGATIVE
  - GCL / SALES_TAX / POA / TOTAL_NET_SALES / NET_CASH: as computed in GL export

Saasant export (credits negative, debits positive) derives its convention by
flipping signs at read time. See `load_frozen_for_saasant`.
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
from pipeline.connection import execute_query_df, execute_sql

# Account name → GL code (reverse of SAASANT_ACCOUNTS in saasant_export.py)
ACCOUNT_TO_GL = {
    "Machine": "401001",
    "Private Pilates": "401002",
    "Class Pass": "401003",
    "Mighty Teacher Training": "401004",
    "Livestream Classes": "401005",
    "Machine Breakage": "403001",
    "Mighty Teacher Training Breakage": "403002",
    "Private Pilates Breakage": "403003",
    "Other Breakage": "403004",
    "Retail Sales": "404000",
    "Refunds": "406000",
    "Discounts": "407000",
}

GL_TO_ACCOUNT = {v: k for k, v in ACCOUNT_TO_GL.items()}

# Saasant sign convention: credits negative, debits positive.
# Revenue accounts are credits, Refunds/Discounts are debits.
SAASANT_DEBIT_ACCOUNTS = {"406000", "407000"}


def is_month_frozen(conn, month_ym: str) -> bool:
    """Return True if this MONTH_YM has any rows in FROZEN_MONTHLY_GL."""
    df = execute_query_df(conn, f"""
        SELECT COUNT(*) AS N
        FROM MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.FROZEN_MONTHLY_GL
        WHERE MONTH_YM = '{month_ym}'
    """)
    return int(df.iloc[0]["N"]) > 0


def load_frozen_gl(conn, month_ym: str) -> pd.DataFrame:
    """
    Load frozen GL for a month, in GL convention.
    Returns columns: MONTH_YM, STUDIO_NAME, LOCATION_NAME, GL_CODE, ACCOUNT, AMOUNT.
    """
    df = execute_query_df(conn, f"""
        SELECT MONTH_YM, STUDIO_NAME, LOCATION_NAME, GL_CODE, ACCOUNT, AMOUNT
        FROM MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.FROZEN_MONTHLY_GL
        WHERE MONTH_YM = '{month_ym}'
        ORDER BY STUDIO_NAME, GL_CODE
    """)
    df["AMOUNT"] = pd.to_numeric(df["AMOUNT"], errors="coerce").astype(float)
    return df


def _rebuild_frozen_gl(conn, month_ym: str, rows: list[dict]) -> None:
    """
    CREATE OR REPLACE pattern (reader-account compatible):
    preserve all rows for other months + replace this month with new rows.
    """
    if not rows:
        values_clause = "SELECT NULL::VARCHAR, NULL::VARCHAR, NULL::VARCHAR, NULL::VARCHAR, NULL::VARCHAR, NULL::NUMBER, NULL::VARCHAR, NULL::VARCHAR, NULL::TIMESTAMP_LTZ WHERE FALSE"
    else:
        parts = []
        for r in rows:
            loc = (r.get("LOCATION_NAME") or "").replace("$", "")
            studio = r["STUDIO_NAME"].replace("$", "")
            acct = r["ACCOUNT"].replace("$", "")
            source = r["SOURCE"].replace("$", "")
            parts.append(
                f"SELECT '{r['MONTH_YM']}', "
                f"$${studio}$$, $${loc}$$, '{r['GL_CODE']}', $${acct}$$, "
                f"{r['AMOUNT']}, '{r['SIGN_CONVENTION']}', $${source}$$, CURRENT_TIMESTAMP()"
            )
        values_clause = " UNION ALL ".join(parts)

    execute_sql(conn, f"""
        CREATE OR REPLACE TABLE MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.FROZEN_MONTHLY_GL AS
        SELECT MONTH_YM, STUDIO_NAME, LOCATION_NAME, GL_CODE, ACCOUNT, AMOUNT, SIGN_CONVENTION, SOURCE, FROZEN_AT
        FROM MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.FROZEN_MONTHLY_GL
        WHERE MONTH_YM != '{month_ym}'
        UNION ALL
        {values_clause}
    """)


def _rebuild_bucket_snapshot(conn, month_ym: str, buckets: dict) -> None:
    if not buckets:
        values_clause = "SELECT NULL::VARCHAR, NULL::VARCHAR, NULL::VARCHAR, NULL::TIMESTAMP_LTZ WHERE FALSE"
    else:
        parts = [
            f"SELECT '{month_ym}', $${st}$$, $${bk}$$, CURRENT_TIMESTAMP()"
            for st, bk in buckets.items()
        ]
        values_clause = " UNION ALL ".join(parts)

    execute_sql(conn, f"""
        CREATE OR REPLACE TABLE MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.SERVICE_TYPE_BUCKETS_FROZEN AS
        SELECT MONTH_YM, SERVICE_TYPE, BUCKET, FROZEN_AT
        FROM MIGHTY_PILATES_ANALYTICS.EARNED_REVENUE_ANALYTICS.SERVICE_TYPE_BUCKETS_FROZEN
        WHERE MONTH_YM != '{month_ym}'
        UNION ALL
        {values_clause}
    """)


def delete_frozen_month(conn, month_ym: str) -> None:
    """Remove all frozen rows for a month (idempotent)."""
    _rebuild_frozen_gl(conn, month_ym, rows=[])
    _rebuild_bucket_snapshot(conn, month_ym, buckets={})


def parse_saasant_file(path: str) -> list[dict]:
    """
    Parse a Saasant JE workbook and extract GL lines.

    Returns list of dicts: MONTH_YM, STUDIO_NAME (full 'Mighty Pilates X'),
    LOCATION_NAME (short 'X'), GL_CODE, ACCOUNT, AMOUNT (GL convention: revenue positive, contra negative).

    Skips the Deferred Revenue balancing row (it's a plug).
    """
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb.active

    # Column indices — read from row 1 headers
    headers = [c.value for c in ws[1]]
    col = {h.strip() if isinstance(h, str) else h: i + 1 for i, h in enumerate(headers)}
    jno_c = col["Journal No"]
    jdate_c = col["Journal Date"]
    acct_c = col["Account"] if "Account" in col else col.get(" Account ") or col[" Account "]
    amt_c = col["Amount"] if "Amount" in col else col.get(" Amount") or col[" Amount"]
    loc_c = col["Location"]

    rows_out = []
    current_location = None
    current_month = None

    for r in range(2, ws.max_row + 1):
        jn = ws.cell(r, jno_c).value
        jdate = ws.cell(r, jdate_c).value
        acct = ws.cell(r, acct_c).value
        amt = ws.cell(r, amt_c).value
        loc = ws.cell(r, loc_c).value

        # Header row: Journal No is a non-formula string like "MB Rev BK 26.01"
        if isinstance(jn, str) and jn.startswith("MB Rev") and not jn.startswith("="):
            current_location = loc
            # Extract month from journal date (preferred) or journal no ("26.01")
            if isinstance(jdate, datetime):
                current_month = jdate.strftime("%Y-%m")
            else:
                parts = jn.split()
                yymm = parts[-1]  # "26.01"
                y, m = yymm.split(".")
                current_month = f"20{y}-{m}"

        if acct is None or amt is None:
            continue
        acct = str(acct).strip()

        # Skip the balancing Deferred Revenue row (it's an Excel SUM formula)
        if acct.replace("  ", " ").lower() == "deferred revenue":
            continue
        # Skip rows with formula amounts (shouldn't happen but guard)
        if isinstance(amt, str):
            continue

        gl_code = ACCOUNT_TO_GL.get(acct)
        if gl_code is None:
            # Unmapped account — log and skip
            print(f"  WARN: unmapped account '{acct}' at row {r}")
            continue

        # Saasant convention: credits negative, debits positive.
        # Flip to GL convention: revenue positive, contra (refunds/discounts) negative.
        if gl_code in SAASANT_DEBIT_ACCOUNTS:
            # Saasant stores debits as positive → GL wants contra as negative
            gl_amount = -abs(amt)
        else:
            # Saasant stores credits as negative → GL wants revenue as positive
            gl_amount = -amt

        rows_out.append({
            "MONTH_YM": current_month,
            "STUDIO_NAME": f"Mighty Pilates {current_location}" if current_location else None,
            "LOCATION_NAME": current_location,
            "GL_CODE": gl_code,
            "ACCOUNT": acct,
            "AMOUNT": round(float(gl_amount), 4),
        })

    return rows_out


def freeze_from_saasant_file(
    conn,
    path: str,
    month_ym: str,
    bucket_dict: dict | None = None,
    force: bool = False,
) -> int:
    """
    Freeze a month by parsing a Saasant JE file and inserting into FROZEN_MONTHLY_GL.

    Returns the number of rows inserted.
    """
    if is_month_frozen(conn, month_ym):
        if not force:
            raise RuntimeError(
                f"Month {month_ym} is already frozen. Pass force=True to overwrite."
            )
        print(f"  Overwriting existing frozen data for {month_ym}...")
        delete_frozen_month(conn, month_ym)

    parsed = parse_saasant_file(path)
    if not parsed:
        raise RuntimeError(f"No rows parsed from {path}")

    # Validate all rows match requested month
    months_seen = {r["MONTH_YM"] for r in parsed}
    if months_seen != {month_ym}:
        raise RuntimeError(
            f"File contains months {months_seen}, expected only {month_ym}"
        )

    for r in parsed:
        r["SIGN_CONVENTION"] = "GL"
        r["SOURCE"] = f"saasant_file:{Path(path).name}"

    _rebuild_frozen_gl(conn, month_ym, parsed)

    if bucket_dict:
        _rebuild_bucket_snapshot(conn, month_ym, bucket_dict)

    print(f"  Froze {len(parsed)} GL rows for {month_ym} from {Path(path).name}")
    return len(parsed)


def load_frozen_for_saasant(conn, month_ym: str) -> pd.DataFrame:
    """
    Load frozen GL in SAASANT convention (credits negative, debits positive).
    Returns columns: MONTH_YM, STUDIO_NAME, LOCATION_NAME, GL_CODE, ACCOUNT, AMOUNT.
    """
    df = load_frozen_gl(conn, month_ym)
    if df.empty:
        return df
    # Flip sign convention: GL (revenue positive, contra negative) → Saasant (credits negative, debits positive)
    def _flip(row):
        if row["GL_CODE"] in SAASANT_DEBIT_ACCOUNTS:
            return abs(row["AMOUNT"])  # debit positive
        return -row["AMOUNT"]  # credit negative
    df = df.copy()
    df["AMOUNT"] = df.apply(_flip, axis=1)
    return df


def load_frozen_for_gl(conn, month_ym: str) -> pd.DataFrame:
    """Load frozen GL in GL convention (as stored). Alias for load_frozen_gl."""
    return load_frozen_gl(conn, month_ym)


def freeze_from_live(conn, year: int, month: int, force: bool = False) -> dict:
    """
    Generate a live Saasant export for the given month, then freeze from it.

    Returns: dict with 'saasant_path' and 'rows_frozen'.
    """
    import calendar
    from pipeline.saasant_export import generate_saasant_export, SERVICE_TYPE_BUCKETS

    last_day = calendar.monthrange(year, month)[1]
    month_ym = f"{year}-{month:02d}"
    start = f"{year}-{month:02d}-01"
    end = f"{year}-{month:02d}-{last_day:02d}"

    if is_month_frozen(conn, month_ym) and not force:
        raise RuntimeError(
            f"Month {month_ym} is already frozen. Pass force=True to re-freeze."
        )

    # Generate live Saasant; skip the frozen fast-path so we compute fresh
    saasant_path = generate_saasant_export(
        conn, start, end,
        output_dir=None,  # default outputs/
        ignore_frozen=True,
    )
    rows = freeze_from_saasant_file(
        conn, saasant_path, month_ym, bucket_dict=SERVICE_TYPE_BUCKETS, force=force,
    )
    return {"saasant_path": saasant_path, "rows_frozen": rows, "month_ym": month_ym}
