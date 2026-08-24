"""
Align the 2025-augmented Excel model's Sales Forecast tab (rows 6-17, cols D-O)
to Cat's authoritative 2025 cash sales CSV.

Cat's CSV represents "Total Sales" (MindBody cash sales) per studio per month
for 2025 — this is the metric the Sales Forecast tab tracks. The current model
was populated from the QBO 2025 P&L "Total Income" (recognized revenue), which
is a related but different metric.

Special handling for Jan-Mar 2025:
  Cat's CSV lumps Presidio + Marin together as "SF" for those months (Marin
  row reads "see above"). We split the SF total into PH and MR using the QBO
  2025 P&L Total Income ratio for the same month.

For studios that don't appear in a month of the CSV, the studio either didn't
exist yet OR had $0 sales. We set those cells to $0.

Output:
  - Updates ~/Desktop/Mighty Pilates/Mighty_Pilates_Financial_Model_2025-2028.xlsx
  - Updates snapshots/excel/Mighty_Pilates_Financial_Model_2025-2028.xlsx
  - Prints a diff report of every change
"""
from __future__ import annotations
import csv
import re
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
CSV_PATH = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Data 23' - 25' - Confidential(2025).csv")
XLSX_DESKTOP = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty_Pilates_Financial_Model_2025-2028.xlsx")
XLSX_SNAPSHOT = REPO / "snapshots" / "excel" / "Mighty_Pilates_Financial_Model_2025-2028.xlsx"
QBO_CONS = REPO / "data" / "financials" / "pl_Dec2025.csv"
QBO_STUDIOS_DIR = REPO / "data" / "financials" / "studios_Dec2025"

MONTHS = ["JANUARY","FEBRUARY","MARCH","APRIL","MAY","JUNE",
          "JULY","AUGUST","SEPTEMBER","OCTOBER","NOVEMBER","DECEMBER"]
MONTH_ABBR = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
MONTH_TO_IDX = {m: i for i, m in enumerate(MONTHS)}

STUDIO_NAME_TO_CODE = {
    "SF": "SF_LUMP",         # special-case marker; resolved later
    "Presidio": "PH",
    "Marin": "MR",
    "SM": "SM",
    "Lafayette": "LF",
    "Berkeley": "BK",
    "Westwood": "WW",
    "Russian Hill": "RH",
    "Ocean Park": "OP",
    "Danville": "DN",
    "Culver City": "CC",
    "West Portal": "WP",
    "Santa Barbara": "SB",
}

# Sales Forecast tab row layout (1-indexed)
ROW_FOR_CODE = {
    "BK": 6, "CC": 7, "DN": 8, "LF": 9, "MR": 10, "OP": 11,
    "PH": 12, "RH": 13, "SB": 14, "SM": 15, "WP": 16, "WW": 17,
}
INSERT_COL = 4  # column D = Jan 2025


def parse_dollar(s: str) -> float:
    """Parse a string like '$250,797 ' into float."""
    if s is None:
        return 0.0
    s = str(s).strip().strip('"').strip()
    if not s:
        return 0.0
    s = s.replace("$", "").replace(",", "").replace(" ", "")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_cat_csv():
    """Parse Cat's CSV into {code: {month_idx: dollars}}.

    SF entries for Jan-Mar (when Marin says 'see above') are stored under the
    SF_LUMP code; later code splits them. SF entries for Apr+ are stored under
    PH (since Marin appears separately then).
    """
    out = {code: {} for code in ROW_FOR_CODE}
    out["SF_LUMP"] = {}  # only for Jan-Mar
    current_month = None
    marin_lumped = False  # tracks whether the current month has Marin as 'see above'

    with open(CSV_PATH) as f:
        for raw in csv.reader(f):
            if not raw:
                continue
            first = (raw[0] or "").strip()
            second = raw[1].strip() if len(raw) > 1 else ""

            # Month header line?
            if first in MONTH_TO_IDX:
                current_month = MONTH_TO_IDX[first]
                marin_lumped = False
                continue

            if current_month is None or not first:
                continue

            # Studio row
            studio_name = first.strip()
            if studio_name not in STUDIO_NAME_TO_CODE:
                # Skip unknown labels (totals row, blank, etc.)
                continue

            # Is this the Marin 'see above' line?
            if studio_name == "Marin" and second.lower().startswith("see above"):
                marin_lumped = True
                continue

            amt = parse_dollar(second)
            if studio_name == "SF":
                if marin_lumped or current_month <= 2:  # Jan-Mar lumped, after that SF=PH
                    out["SF_LUMP"][current_month] = amt
                else:
                    out["PH"][current_month] = amt
                continue

            code = STUDIO_NAME_TO_CODE[studio_name]
            out[code][current_month] = amt

    return out


def split_sf_lump(cat: dict) -> dict:
    """For months where SF was lumped (Jan-Mar), split into PH and MR using
    QBO 2025 P&L Total Income ratios."""
    ph_csv = pd.read_csv(QBO_STUDIOS_DIR / "PH_Dec2025.csv").set_index("Account")
    mr_csv = pd.read_csv(QBO_STUDIOS_DIR / "MR_Dec2025.csv").set_index("Account")

    if "Total Income" not in ph_csv.index or "Total Income" not in mr_csv.index:
        raise RuntimeError("Could not find 'Total Income' in QBO PH/MR CSV")

    for month_idx, sf_total in list(cat.get("SF_LUMP", {}).items()):
        month_label = f"{MONTH_ABBR[month_idx]} 2025"
        ph_qbo = float(ph_csv.loc["Total Income", month_label] or 0)
        mr_qbo = float(mr_csv.loc["Total Income", month_label] or 0)
        denom = ph_qbo + mr_qbo
        if denom <= 0:
            # Edge case — put it all in PH
            ph_share, mr_share = 1.0, 0.0
        else:
            ph_share, mr_share = ph_qbo / denom, mr_qbo / denom
        cat["PH"][month_idx] = sf_total * ph_share
        cat["MR"][month_idx] = sf_total * mr_share
        print(f"   Split SF for {month_label}: ${sf_total:,.0f} → "
              f"PH ${sf_total*ph_share:,.0f} ({ph_share*100:.1f}%), "
              f"MR ${sf_total*mr_share:,.0f} ({mr_share*100:.1f}%)")
    cat.pop("SF_LUMP", None)
    return cat


def update_excel(cat: dict):
    """Update Sales Forecast tab rows 6-17 cols D-O to match cat data.
    Studios absent from a given month get $0 (didn't exist yet)."""
    wb = load_workbook(XLSX_DESKTOP)
    ws = wb["Sales Forecast"]

    print(f"\n=== Diff report (Sales Forecast Jan-Dec 2025) ===")
    print(f"{'Studio':<6} {'Month':<10} {'Was':>14} {'Now':>14} {'Δ':>14}")
    print("-" * 60)

    total_changes = 0
    total_old = total_new = 0
    for code, row in ROW_FOR_CODE.items():
        for m_idx in range(12):
            new_val = float(cat.get(code, {}).get(m_idx, 0.0))
            col = INSERT_COL + m_idx
            cell = ws.cell(row=row, column=col)
            old_val = cell.value
            old_num = float(old_val) if isinstance(old_val, (int, float)) else 0.0
            total_old += old_num
            total_new += new_val
            if abs(old_num - new_val) > 0.5:
                print(f"{code:<6} {MONTH_ABBR[m_idx]+' 2025':<10} "
                      f"${old_num:>12,.0f} ${new_val:>12,.0f} ${new_val-old_num:>+12,.0f}")
                total_changes += 1
            cell.value = new_val

    print("-" * 60)
    print(f"Total cells changed: {total_changes}")
    print(f"2025 total — old: ${total_old:,.0f}  new: ${total_new:,.0f}  Δ: ${total_new-total_old:+,.0f}")

    wb.save(XLSX_DESKTOP)
    shutil.copy(XLSX_DESKTOP, XLSX_SNAPSHOT)
    print(f"\nSaved: {XLSX_DESKTOP}")
    print(f"        {XLSX_SNAPSHOT.relative_to(REPO)}")


def main():
    print("Parsing Cat's CSV...")
    cat = parse_cat_csv()

    print("\nSplitting Jan-Mar 'SF' lump into PH + MR by QBO ratio:")
    cat = split_sf_lump(cat)

    print("\nCSV totals by studio code (2025 annual):")
    grand = 0
    for code in sorted(ROW_FOR_CODE):
        total = sum(cat.get(code, {}).values())
        grand += total
        print(f"  {code}: ${total:>14,.0f}")
    print(f"  {'TOTAL':<3}: ${grand:>14,.0f}")

    update_excel(cat)


if __name__ == "__main__":
    main()
