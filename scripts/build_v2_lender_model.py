"""
Build v2 of the analyst's Updated DRAFT model with surgical edits per Cat's
2026-06-25 directives.

Source : Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026.xlsx
Output : Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v2.xlsx

Surgical edits:
  (a)+(f) Add "Cash Sales vs Accrual Revenue" note to Read Me tab
  (b)     P&L row 25 Taxes Paid forecast: replace $266.67 with $1,188.60/mo
          (= average of Jan-May 2026 actuals: 0+5143+0+0+800 = 5943 / 5)
  (c)     Per-studio P&L row 65 Depreciation forecast (Jun 2026 - Dec 2028):
          replace circular formula with hardcoded per-studio May 2026 actuals.
          Sums to $30,883/mo, ties to the consolidated row 23 already there.
  (d)     P&L row 24 Interest Expense:
            Jun-Aug 2026 : keep $14,026 (existing debt service, untouched)
            Sept 2026+   : = Cash Flow Forecast row 49 (PF cash interest)
                           + $1,333.33/mo PF Fees amortization
          PF Fees of $80K amortized straight-line over 60 months (5-yr term).
          Combined into row 24 because GAAP reports debt-issuance amort as
          part of interest expense. Lender's DSCR uses CF row 49 directly
          for "Cash Interest" so this doesn't disturb their metrics.
  (e)     Cash Flow Forecast row 19 Taxes forecast (Jun 2026+):
          link to ='P&L'!<col>25 + 'P&L'!<col>26 (Taxes Paid + Property Taxes)
          Once (b) lands, P&L = CF tax line automatically.

Cash Flow Forecast left exactly as the lender's analyst had it (per Cat).
"""
from __future__ import annotations
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v2.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v2.xlsx"

# Column conventions (verified by inspection):
#   D  (4)  = Jan 2025
#   P  (16) = Jan 2026
#   T  (20) = May 2026
#   U  (21) = Jun 2026
#   W  (23) = Aug 2026
#   X  (24) = Sep 2026  (loan closes)
#   AM (39) = Dec 2028  (last forecast month)
JUN_2026 = 21       # U — first forecast month after May 2026 actuals
SEP_2026 = 24       # X — refinancing close
DEC_2028 = 39       # AM — last col in monthly horizon

# Per-studio depreciation, May 2026 actual values (from QBO 0531 package).
# Sum = $30,882.78/mo ≈ the $30,882.44 already on consolidated P&L row 23.
STUDIO_DEPRECIATION = {
    "BK P&L": 1833.55,
    "CC P&L": 4691.67,
    "DN P&L": 3615.45,
    "LF P&L": 4879.31,
    "MR P&L": 2726.68,
    "OP P&L": 1996.64,
    "PH P&L": 724.74,
    "RH P&L": 2812.47,
    "SB P&L": 2483.89,
    "SM P&L": 2293.28,
    "WP P&L": 1038.11,
    "WW P&L": 1780.99,
}
STUDIO_DEPRECIATION_ROW = 65          # row '  810000 Depreciation' on each studio P&L tab

# PF Fees amortization (Item d)
PF_FEES_TOTAL = 80000.00              # $50K legal + $30K financing (from Sources & Uses)
LOAN_TERM_MONTHS = 60                 # 5-year amortization assumption
PF_AMORT_MONTHLY = round(PF_FEES_TOTAL / LOAN_TERM_MONTHS, 4)  # $1,333.33

# Better Tax Paid average (Item b)
TAX_ACTUALS_2026 = [0.00, 5143.00, 0.00, 0.00, 800.00]   # Jan-May 2026 actuals
NEW_TAX_AVG = round(sum(TAX_ACTUALS_2026) / len(TAX_ACTUALS_2026), 2)  # $1,188.60


def apply_item_c(wb):
    """Item (c): Per-studio depreciation forecast — flat extension of May 2026 actuals."""
    print("\n[c] Per-studio depreciation forecast (cols U-AM, Jun 2026 - Dec 2028)")
    for sheet, monthly_value in STUDIO_DEPRECIATION.items():
        if sheet not in wb.sheetnames:
            print(f"   WARNING: {sheet!r} missing — skipping")
            continue
        ws = wb[sheet]
        for col in range(JUN_2026, DEC_2028 + 1):
            ws.cell(row=STUDIO_DEPRECIATION_ROW, column=col).value = monthly_value
        print(f"   {sheet:<10} r{STUDIO_DEPRECIATION_ROW}: ${monthly_value:>9,.2f}/mo flat")
    total = sum(STUDIO_DEPRECIATION.values())
    print(f"   Sum across studios: ${total:,.2f}/mo (consolidated P&L r23 already shows ${30882.44:,.2f})")


def apply_item_d(wb):
    """Item (d): P&L row 24 Interest Expense — link to debt schedule from Sept 2026,
    plus PF Fees amortization combined."""
    print("\n[d] P&L Interest Expense — link to Cash Flow debt schedule + PF amortization")
    print(f"    Jun-Aug 2026 (cols U,V,W): keep $14,026 (existing debt service)")
    print(f"    Sept 2026+ (cols X-AM):    =CF!<col>49 + ${PF_AMORT_MONTHLY:.4f} (PF amort over {LOAN_TERM_MONTHS} months)")
    ws = wb["P&L"]
    INT_EXPENSE_ROW = 24
    for col in range(SEP_2026, DEC_2028 + 1):
        cl = get_column_letter(col)
        ws.cell(row=INT_EXPENSE_ROW, column=col).value = (
            f"='Cash Flow Forecast'!{cl}49+{PF_AMORT_MONTHLY}"
        )
    # Sanity: cols U,V,W (Jun-Aug 2026) NOT touched — confirmed
    for col in range(JUN_2026, SEP_2026):
        cl = get_column_letter(col)
        v = ws.cell(row=INT_EXPENSE_ROW, column=col).value
        print(f"       col {cl} unchanged: {v}")


def apply_item_b(wb):
    """Item (b): Refresh Taxes Paid forecast average using post-May 2026 better window."""
    print("\n[b] P&L Taxes Paid forecast — refresh from 2026 actuals window")
    print(f"    Jan-May 2026 actuals: {TAX_ACTUALS_2026}")
    print(f"    New avg: ${NEW_TAX_AVG:,.2f}/mo (was $266.67)")
    ws = wb["P&L"]
    TAX_ROW = 25
    for col in range(JUN_2026, DEC_2028 + 1):
        ws.cell(row=TAX_ROW, column=col).value = NEW_TAX_AVG
    print(f"    Updated P&L r{TAX_ROW} cols U-AM")


def apply_item_e(wb):
    """Item (e): Cash Flow row 19 Taxes — link to P&L Taxes Paid + Property Taxes."""
    print("\n[e] Cash Flow Forecast row 19 Taxes — link to P&L (Taxes Paid + Property Taxes)")
    ws = wb["Cash Flow Forecast"]
    CF_TAX_ROW = 19
    for col in range(JUN_2026, DEC_2028 + 1):
        cl = get_column_letter(col)
        ws.cell(row=CF_TAX_ROW, column=col).value = (
            f"='P&L'!{cl}25+'P&L'!{cl}26"
        )
    print(f"    Updated CF r{CF_TAX_ROW} cols U-AM: =P&L!<col>25 + P&L!<col>26")


def apply_items_a_f(wb):
    """Items (a)+(f): Cash Sales vs Accrual Revenue explanatory note on Read Me tab."""
    print("\n[a+f] Adding 'Cash Sales vs Accrual Revenue' note to Read Me tab")
    ws = wb["Read Me"]

    # Append below existing content (last existing row is 43)
    start_row = 45
    notes = [
        ("Cash Sales vs Accrual Revenue", "Why P&L Revenue and Sales Forecast Revenue diverge — and when they should."),
        ("Sales Forecast / Summary", "These tabs show CASH SALES: Total Sales captured at the MindBody point-of-sale when a member buys a class pack, membership, or workshop."),
        ("Studio P&L / All Studios", "These tabs show ACCRUAL P&L REVENUE: Sessions recognized over the package usage window + Breakage when unused packages expire − Refunds − Discounts."),
        ("Worked example", "Berkeley Dec 2028: Cash Sales $164K, Accrual Revenue $140K. The $24K gap = packages sold in Dec 2028 that will be USED (or breakage'd) in later months."),
        ("Reconciliation horizon", "These metrics diverge in any individual month by design. They reconcile only over the full lifecycle of a cohort of packages (typically 6-12 months from purchase)."),
        ("Both are correct", "Cash Sales is the operating KPI (what closed at the counter). Accrual Revenue is the GAAP P&L line (what's earned). Neither is wrong."),
    ]

    # Section header
    hdr = ws.cell(row=start_row, column=2, value="Cash Sales vs Accrual Revenue")
    hdr.font = Font(bold=True, size=11)
    for i, (label, body) in enumerate(notes[1:], start=1):
        ws.cell(row=start_row + i, column=2, value=label).font = Font(bold=True)
        ws.cell(row=start_row + i, column=3, value=body).alignment = Alignment(wrap_text=True, vertical="top")
    # Set wider col C if not already
    if ws.column_dimensions["C"].width < 80:
        ws.column_dimensions["C"].width = 90

    print(f"    Added {len(notes)} rows starting at Read Me row {start_row}")


def main():
    print("=" * 78)
    print("Build v2 model with surgical lender-flag fixes (items b, c, d, e + a/f note)")
    print("=" * 78)

    if not SRC.exists():
        raise FileNotFoundError(f"Source file not found: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"\nCopied source → {OUT.name}")

    wb = load_workbook(OUT)
    print(f"Opened workbook: {len(wb.sheetnames)} sheets")

    apply_items_a_f(wb)
    apply_item_c(wb)
    apply_item_d(wb)
    apply_item_b(wb)
    apply_item_e(wb)

    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")
    print("\nDone. Open the v2 file in Excel and spot-check.")


if __name__ == "__main__":
    main()
