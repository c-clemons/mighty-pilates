"""
Build a SEPARATE Excel model that extends the live 2026 model with 2025 actuals.

Approach (per Cat 2026-06-23):
  - Insert 12 columns (Jan 2025 - Dec 2025) before Jan 2026 in all month-based sheets.
  - Rewrite every formula in the workbook so column references that point to
    expanding sheets at col >= D shift by +12. Refs to non-expanding sheets
    (Assumptions, Cover, etc.) stay put.
  - Populate the new 2025 columns with hard-coded actuals from the accountant's
    2025 financial package (already imported to data/financials/*Dec2025.csv).
  - Save as a NEW file — does not touch the live model.

Output: Mighty_Pilates_Financial_Model_2025-2028.xlsx (in ~/Desktop/Mighty Pilates/
        and in snapshots/excel/ for repo tracking).
"""
from __future__ import annotations
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

# ------------------------------------------------------------------
# Config
# ------------------------------------------------------------------
REPO = Path(__file__).resolve().parents[1]
LIVE = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Pilates Financial Model.xlsx")
DESKTOP_OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty_Pilates_Financial_Model_2025-2028.xlsx")
SNAPSHOT_OUT = REPO / "snapshots" / "excel" / "Mighty_Pilates_Financial_Model_2025-2028.xlsx"

FINANCIALS = REPO / "data" / "financials"
CONS_CSV = FINANCIALS / "pl_Dec2025.csv"
STUDIO_CSV_DIR = FINANCIALS / "studios_Dec2025"

INSERT_COL = 4           # column D
N_NEW_COLS = 12          # Jan-Dec 2025
SHIFT_THRESHOLD = INSERT_COL  # any col index >= this in expanding sheets shifts by +N_NEW_COLS

EXPANDING_SHEETS = {
    "Summary", "Sales Forecast", "Cash, Debt & Equity",
    "Cash Flow Forecast", "P&L", "All Studios Summary",
    "BK P&L", "CC P&L", "DN P&L", "LF P&L", "MR P&L",
    "OP P&L", "PH P&L", "RH P&L", "SB P&L", "SM P&L",
    "WP P&L", "WW P&L", "HO P&L",
}

# Studios that have a dedicated P&L tab in the Excel model
STUDIO_TABS = {
    "BK": "Berkeley", "CC": "Culver City", "DN": "Danville", "LF": "Lafayette",
    "MR": "Marin", "OP": "Ocean Park", "PH": "Presidio Heights",
    "RH": "Russian Hill", "SB": "Santa Barbara", "SM": "Santa Monica",
    "WP": "West Portal", "WW": "Westwood", "HO": "Head Office",
}

MONTH_LABELS = [f"{m} 2025" for m in
                ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]]


# ------------------------------------------------------------------
# Formula rewriting
# ------------------------------------------------------------------
# Matches optional sheet prefix + cell reference.
# Sheet name: either 'Quoted Name' or unquoted identifier (letters/digits/underscore/period).
# Cell: optional $, col letters, optional $, digits.
REF_RE = re.compile(
    r"""(?:(?P<sheet>'(?:[^']|'')+'|[A-Za-z_][\w.]*)!)?
        (?P<col_abs>\$?)(?P<col>[A-Z]{1,3})(?P<row_abs>\$?)(?P<row>\d+)""",
    re.VERBOSE,
)


def _shift_col(col_letter: str) -> str:
    idx = column_index_from_string(col_letter)
    if idx < SHIFT_THRESHOLD:
        return col_letter
    return get_column_letter(idx + N_NEW_COLS)


def rewrite_formula(formula: str, current_sheet: str) -> str:
    """Rewrite column references in a formula based on expansion rules."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def _sub(m: re.Match) -> str:
        sheet_token = m.group("sheet")
        col_abs = m.group("col_abs")
        col = m.group("col")
        row_abs = m.group("row_abs")
        row = m.group("row")

        if sheet_token:
            sheet_name = sheet_token.strip("'").replace("''", "'")
        else:
            sheet_name = current_sheet

        if sheet_name not in EXPANDING_SHEETS:
            return m.group(0)

        new_col = _shift_col(col)
        sheet_part = f"{sheet_token}!" if sheet_token else ""
        return f"{sheet_part}{col_abs}{new_col}{row_abs}{row}"

    return REF_RE.sub(_sub, formula)


def rewrite_all_formulas(wb):
    """Walk every cell on every sheet and rewrite formulas."""
    cell_count = 0
    rewritten_count = 0
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_count += 1
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    new = rewrite_formula(cell.value, sname)
                    if new != cell.value:
                        cell.value = new
                        rewritten_count += 1
    print(f"  Walked {cell_count} cells; rewrote {rewritten_count} formulas")


# ------------------------------------------------------------------
# Column insertion + data population
# ------------------------------------------------------------------
def insert_columns_in_expanding_sheets(wb):
    for sname in EXPANDING_SHEETS:
        if sname not in wb.sheetnames:
            print(f"  WARNING: expanding sheet {sname!r} not found, skipping")
            continue
        ws = wb[sname]
        ws.insert_cols(INSERT_COL, amount=N_NEW_COLS)
        print(f"  Inserted {N_NEW_COLS} cols at col D in {sname}")


def write_month_headers(wb):
    """Populate Jan 2025...Dec 2025 in the new row-5 header cells."""
    for sname in EXPANDING_SHEETS:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        # Some sheets use row 5 for column headers; HO P&L might use "(A)" suffix.
        # Detect header row by finding the row that has "Jan 2026" or "Feb 2026" at col P (post-insert).
        # Simpler: just write to row 5 since that's the convention.
        for i, label in enumerate(MONTH_LABELS):
            cell = ws.cell(row=5, column=INSERT_COL + i)
            # Preserve "(A)" suffix style if Jan 2026 (now at col P) has it
            jan26_label = ws.cell(row=5, column=INSERT_COL + N_NEW_COLS).value
            if isinstance(jan26_label, str) and jan26_label.endswith("(A)"):
                cell.value = f"{label} (A)"
            else:
                cell.value = label


def load_2025_data():
    """Read 2025 CSVs into dicts keyed by (studio_code or 'CONSOLIDATED', account, month)."""
    data = {"CONSOLIDATED": {}}
    cons_df = pd.read_csv(CONS_CSV).set_index("Account")
    months = [c for c in cons_df.columns if c != "Total"]
    for acct in cons_df.index:
        data["CONSOLIDATED"][acct] = {m: float(cons_df.loc[acct, m] or 0) for m in months}

    for csv in sorted(STUDIO_CSV_DIR.glob("*_Dec2025.csv")):
        code = csv.stem.split("_Dec2025")[0]
        df = pd.read_csv(csv).set_index("Account")
        data[code] = {}
        for acct in df.index:
            data[code][acct] = {m: float(df.loc[acct, m] or 0) for m in df.columns if m != "Total"}
    return data


def get_account_label(ws, row: int) -> str | None:
    """Look up the account label from col B of the given row."""
    label = ws.cell(row=row, column=2).value
    if label is None:
        return None
    return str(label).strip().lstrip("  ").strip()


# Explicit map: model row label → CSV row label (for subtotals that don't share
# the same text). Leaf accounts match by direct label comparison.
# Model labels are stored without leading whitespace (we strip in get_account_label).
MODEL_LABEL_TO_CSV = {
    # Studio P&L subtotals
    "Total Sessions":                  "Total 401000 Sessions",
    "Total Breakage Revenue":          "Total 403000 Breakage Revenue",
    "TOTAL REVENUE":                   "Total Income",
    "Total Cost of Goods Sold":        "Total Cost of Goods Sold",
    "GROSS PROFIT":                    "Gross Profit",
    "Total Marketing":                 "Total 601000 Sales & Marketing",
    "Total Payroll":                   "Total 602000 Payroll",
    "Total Professional Fees":         "Total 604000 Professional Fees",
    "Total Utilities":                 "Total 616000 Utilities",
    "Total Property Costs":            "Total 700000 Property Costs",
    "TOTAL OPERATING EXPENSES":        "Total Expenses",
    "NET OPERATING INCOME (EBITDA)":   "Net Operating Income",
    "Total Other Expenses":            "Total Other Expenses",
    "NET INCOME":                      "Net Income",

    # Consolidated P&L tab (different label conventions)
    "Session Revenue":                 "Total 401000 Sessions",
    "Breakage Revenue":                "Total 403000 Breakage Revenue",
    "Retail Sales":                    "404000 Retail Sales",
    "Refunds":                         "406000 Refunds",
    "Discounts":                       "407000 Discounts",
    "Total Revenue":                   "Total Income",
    "Gross Profit":                    "Gross Profit",
    "Marketing":                       "Total 601000 Sales & Marketing",
    "Payroll":                         "Total 602000 Payroll",
    "Software":                        "603000 Software & Web Services",
    "Professional Fees":               "Total 604000 Professional Fees",
    "Utilities":                       "Total 616000 Utilities",
    "Property Costs":                  "Total 700000 Property Costs",
    "Total Operating Expenses":        "Total Expenses",
    "Net Operating Income":            "Net Operating Income",
    "Depreciation":                    "810000 Depreciation",
    "Taxes Paid":                      "902000 Taxes Paid",
    "Property Taxes":                  "903000 Property taxes",
    "Net Income":                      "Net Income",
}

# Model rows whose subtotal doesn't exist in the CSV — must be computed by
# summing the immediately-preceding leaf rows. Key = model label, value = list
# of leaf-row offsets (relative to subtotal row).
MODEL_SUBTOTAL_FORMULAS = {
    # "Total Software & Admin" at row 42 sums rows 39 (603000), 40 (608000), 41 (610000)
    "Total Software & Admin":  [-3, -2, -1],
    # "Total Travel & Meals" at row 50 sums rows 48 (605000), 49 (606000)
    "Total Travel & Meals":    [-2, -1],
}


def populate_studio_pl(ws, studio_data: dict):
    """Populate per-studio P&L 2025 cols (D-O) by matching row labels.
    Leaf rows match exact label; subtotal rows use MODEL_LABEL_TO_CSV; subtotals
    without CSV equivalents get computed SUM formulas."""
    populated = 0
    formula_subtotal = 0
    unmatched = []

    csv_keys_lc = {k.strip().lower(): k for k in studio_data.keys()}

    for row in range(6, ws.max_row + 1):
        label = get_account_label(ws, row)
        if not label:
            continue
        canon = label.strip()

        # Try MODEL_LABEL_TO_CSV first
        csv_label = MODEL_LABEL_TO_CSV.get(canon)
        if csv_label is None:
            # Try direct match (case-insensitive)
            csv_label = csv_keys_lc.get(canon.lower())

        if csv_label and csv_label in studio_data:
            monthly = studio_data[csv_label]
            for i, m in enumerate(MONTH_LABELS):
                val = monthly.get(m, 0) or 0
                ws.cell(row=row, column=INSERT_COL + i).value = float(val)
            populated += 1
        elif canon in MODEL_SUBTOTAL_FORMULAS:
            # Use computed SUM of preceding leaves
            offsets = MODEL_SUBTOTAL_FORMULAS[canon]
            for i in range(N_NEW_COLS):
                col_letter = get_column_letter(INSERT_COL + i)
                parts = [f"{col_letter}{row + off}" for off in offsets]
                ws.cell(row=row, column=INSERT_COL + i).value = "=" + "+".join(parts)
            formula_subtotal += 1
        else:
            # Header rows or unmapped — leave blank
            if not canon.isupper() and not canon.endswith("EXPENSES") and canon not in {
                "INCOME","COST OF GOODS SOLD","OPERATING EXPENSES","OTHER",
            }:
                unmatched.append(canon)

    return populated, formula_subtotal, unmatched


def populate_sales_forecast_2025(ws, by_studio: dict):
    """Populate Sales Forecast tab rows 6-17 cols D-O with each studio's 2025
    Total Income per month."""
    code_to_row = {}
    for r in range(6, 18):
        code = ws.cell(row=r, column=2).value
        if code:
            code_to_row[str(code).strip()] = r

    populated = 0
    for code, row in code_to_row.items():
        sd = by_studio.get(code, {})
        total_income = sd.get("Total Income", {})
        if not total_income:
            continue
        for i, m in enumerate(MONTH_LABELS):
            v = total_income.get(m, 0) or 0
            ws.cell(row=row, column=INSERT_COL + i).value = float(v)
        populated += 1
    return populated


def add_total_row_formulas(ws):
    """For Sales Forecast row 18 (TOTAL): add =SUM(D6:D17) ... for each new 2025 column."""
    for i in range(N_NEW_COLS):
        col_letter = get_column_letter(INSERT_COL + i)
        ws.cell(row=18, column=INSERT_COL + i).value = f"=SUM({col_letter}6:{col_letter}17)"


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------
def main():
    print(f"\n{'='*78}")
    print(f"BUILD 2025-augmented Excel model")
    print(f"{'='*78}\n")

    # 1. Copy live workbook to new file
    DESKTOP_OUT.parent.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_OUT.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(LIVE, DESKTOP_OUT)
    print(f"[1] Copied live model → {DESKTOP_OUT.name}")

    # 2. Open with openpyxl
    wb = load_workbook(DESKTOP_OUT)
    print(f"[2] Opened workbook: {len(wb.sheetnames)} sheets")

    # 3. Rewrite formulas (BEFORE inserting columns, so refs match post-insertion positions)
    print(f"\n[3] Rewriting formulas to shift col refs +{N_NEW_COLS} in expanding sheets")
    rewrite_all_formulas(wb)

    # 4. Insert 12 columns at col D in expanding sheets
    print(f"\n[4] Inserting {N_NEW_COLS} columns at col D in {len(EXPANDING_SHEETS)} sheets")
    insert_columns_in_expanding_sheets(wb)

    # 5. Write month headers (Jan 2025 ... Dec 2025) at row 5 of new cols
    print(f"\n[5] Writing month headers in new D-O cells")
    write_month_headers(wb)

    # 6. Load 2025 data + populate per-studio + consolidated P&L
    print(f"\n[6] Populating 2025 actuals from CSVs")
    data = load_2025_data()
    print(f"  Loaded data for: {sorted(data.keys())}")

    for code, sheet_name in STUDIO_TABS.items():
        tab = f"{code} P&L"
        if tab not in wb.sheetnames:
            print(f"  WARNING: sheet {tab!r} missing — skipping")
            continue
        if code not in data:
            print(f"  WARNING: no 2025 data for {code} — skipping")
            continue
        n, nf, unm = populate_studio_pl(wb[tab], data[code])
        print(f"  {tab}: {n} rows populated, {nf} subtotal formulas added"
              + (f", {len(unm)} unmatched: {unm[:3]}..." if unm else ""))

    if "P&L" in wb.sheetnames:
        n, nf, unm = populate_studio_pl(wb["P&L"], data["CONSOLIDATED"])
        print(f"  P&L (consolidated): {n} rows populated, {nf} formulas"
              + (f", unmatched: {unm[:3]}" if unm else ""))

    # 7. Sales Forecast — per-studio Total Income for 2025
    if "Sales Forecast" in wb.sheetnames:
        n = populate_sales_forecast_2025(wb["Sales Forecast"], data)
        add_total_row_formulas(wb["Sales Forecast"])
        print(f"  Sales Forecast: populated {n} studio rows + TOTAL formulas")

    # 8. Save
    wb.save(DESKTOP_OUT)
    print(f"\n[7] Saved: {DESKTOP_OUT}")
    shutil.copy(DESKTOP_OUT, SNAPSHOT_OUT)
    print(f"        Snapshot: {SNAPSHOT_OUT.relative_to(REPO)}")
    print(f"\nDone.")


if __name__ == "__main__":
    main()
