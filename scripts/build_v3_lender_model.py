"""
v3 of the analyst's lender model.

v2 → v3 changes per Cat's 2026-06-30 directives (no other edits):

Consolidated P&L tab — replace hardcoded forecast values with SUM links to
individual P&L tabs (studio P&Ls + HO P&L). Forecast cols only (Jun 2026 →
Dec 2028, cols U-AM). Actuals (Jan 2025 → May 2026) untouched.

  r8  Retail Sales      → SUM('BK P&L'!<col>18 + ... + 'WW P&L'!<col>18)
  r23 Depreciation      → SUM(studios r65 + 'HO P&L'!<col>38)
  r24 Interest Expense  → SUM(studios r66 + 'HO P&L'!<col>39)
                         [recap interest formula moved from P&L r24 to HO r39]
  r25 Taxes Paid        → SUM(studios r67 + 'HO P&L'!<col>33)
                         [studio r67 forecast zeroed; HO r33 set to $1,188.60]

Cash Flow Forecast — replace hardcoded forecast values with links to
consolidated P&L. Cat: "accrual expenses = cash expenses." Forecast cols
only (Jun 2026 → Dec 2028, cols U-AM).

  r10 Property Costs       → ='P&L'!<col>19
  r11 Staff Costs          → ='P&L'!<col>15
  r12 Utilities            → ='P&L'!<col>18
  r13 Marketing & Promotion → ='P&L'!<col>14
  r14 Administrative & G&A → ='P&L'!<col>16  (Software & Admin)
  r15 Professional Fees    → ='P&L'!<col>17
  r16 Travel & Meals       → SUM(studio r50 + HO r30) — consolidated has no row for this
  r17 Merchant Fees & COGS → ='P&L'!<col>12
  r18 Studio Start Up Costs → unchanged (=0)
  r19 Taxes                → already linked in v2 to P&L r25+r26 (unchanged)

NOT changed (worth surfacing to Cat):
  - Consolidated P&L r26 Property Taxes: kept hardcoded $3,673. No clean per-
    studio source — studios' row 68 is $0 in actuals; HO P&L has no property
    tax row. Would require structural row insert on HO P&L to link cleanly.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v2.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v3.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v3.xlsx"

JUN_2026 = 21       # U
SEP_2026 = 24       # X
DEC_2028 = 39       # AM

STUDIO_TABS = ['BK P&L', 'CC P&L', 'DN P&L', 'LF P&L', 'MR P&L', 'OP P&L',
               'PH P&L', 'RH P&L', 'SB P&L', 'SM P&L', 'WP P&L', 'WW P&L']

# Studio P&L rows
STUDIO_RETAIL_ROW = 18           # 404000 Retail Sales
STUDIO_DEPRECIATION_ROW = 65     # 810000 Depreciation
STUDIO_INTEREST_ROW = 66         # 901000 Interest Expense
STUDIO_TAXES_ROW = 67            # 902000 Taxes Paid
STUDIO_TRAVEL_MEALS_TOTAL_ROW = 50  # Total Travel & Meals subtotal

# HO P&L rows
HO_TAXES_ROW = 33                # 902000 Taxes (Franchise/State)
HO_DEPRECIATION_ROW = 38         # 810000 Depreciation
HO_INTEREST_ROW = 39             # 901000 Interest Expense
HO_TRAVEL_MEALS_TOTAL_ROW = 30   # Total Travel & Meals subtotal

# Consolidated P&L rows
PL_RETAIL_ROW = 8
PL_DEPRECIATION_ROW = 23
PL_INTEREST_ROW = 24
PL_TAXES_ROW = 25
PL_PROPERTY_TAXES_ROW = 26
# Categories the CF tab needs to reach:
PL_MARKETING_ROW = 14
PL_PAYROLL_ROW = 15
PL_SOFTWARE_ROW = 16            # = Software & Admin
PL_PROFESSIONAL_FEES_ROW = 17
PL_UTILITIES_ROW = 18
PL_PROPERTY_COSTS_ROW = 19
PL_COGS_ROW = 12                # COGS & Merchant Fees

# Cash Flow Forecast rows
CF_PROPERTY_COSTS_ROW = 10
CF_STAFF_COSTS_ROW = 11
CF_UTILITIES_ROW = 12
CF_MARKETING_ROW = 13
CF_ADMIN_ROW = 14
CF_PROFESSIONAL_FEES_ROW = 15
CF_TRAVEL_MEALS_ROW = 16
CF_MERCHANT_COGS_ROW = 17

# Item (b) entity-level operating tax target value
ENTITY_OPERATING_TAX_MO = 1188.60

# PF interest amortization (from v2)
PF_AMORT_MONTHLY = round(80000.00 / 60, 4)


def studio_sum(col_letter: str, row: int, include_ho: bool = False, ho_row: int | None = None) -> str:
    parts = [f"'{tab}'!{col_letter}{row}" for tab in STUDIO_TABS]
    if include_ho and ho_row:
        parts.append(f"'HO P&L'!{col_letter}{ho_row}")
    return "=" + "+".join(parts)


def move_interest_formula_to_ho(wb):
    """Step 1: move the recap interest formula from P&L r24 to HO r39.
    Jun-Aug 2026: keep $14,026 on HO r39. Sept 2026+: CF row 49 + amortization.
    P&L r24 will then SUM(studio r66 + HO r39) → equals HO r39 since studios are 0.
    """
    print("\n[1] Move recap interest formula from P&L r24 → HO P&L r39")
    ho = wb["HO P&L"]
    for col in range(JUN_2026, SEP_2026):
        ho.cell(row=HO_INTEREST_ROW, column=col).value = 14026
    for col in range(SEP_2026, DEC_2028 + 1):
        cl = get_column_letter(col)
        ho.cell(row=HO_INTEREST_ROW, column=col).value = f"='Cash Flow Forecast'!{cl}49+{PF_AMORT_MONTHLY}"
    print(f"    HO r{HO_INTEREST_ROW} U,V,W = 14026; X-AM = CF!<col>49 + {PF_AMORT_MONTHLY}")


def update_ho_operating_tax(wb):
    """Step 2: HO r33 forecast → $1,188.60/mo (was $5,387.33)."""
    print("\n[2] HO r33 Taxes (Franchise/State) forecast → $1,188.60/mo")
    ho = wb["HO P&L"]
    for col in range(JUN_2026, DEC_2028 + 1):
        ho.cell(row=HO_TAXES_ROW, column=col).value = ENTITY_OPERATING_TAX_MO


def zero_studio_taxes_forecast(wb):
    """Step 3: studio P&L r67 forecast → 0 (so HO carries entity operating tax)."""
    print("\n[3] Studio P&L r67 Taxes Paid forecast → 0 across all 12 studios")
    for tab in STUDIO_TABS:
        ws = wb[tab]
        for col in range(JUN_2026, DEC_2028 + 1):
            ws.cell(row=STUDIO_TAXES_ROW, column=col).value = 0


def link_consolidated_pl(wb):
    """Step 4: rewire consolidated P&L forecast rows to SUM(individual P&Ls)."""
    print("\n[4] Consolidated P&L forecast → SUM(individual P&L tabs)")
    pl = wb["P&L"]
    for col in range(JUN_2026, DEC_2028 + 1):
        cl = get_column_letter(col)
        # Retail Sales — studios only (HO has no retail)
        pl.cell(row=PL_RETAIL_ROW, column=col).value = studio_sum(cl, STUDIO_RETAIL_ROW)
        # Depreciation — studios + HO
        pl.cell(row=PL_DEPRECIATION_ROW, column=col).value = studio_sum(cl, STUDIO_DEPRECIATION_ROW, True, HO_DEPRECIATION_ROW)
        # Interest Expense — studios + HO (HO now carries the recap formula)
        pl.cell(row=PL_INTEREST_ROW, column=col).value = studio_sum(cl, STUDIO_INTEREST_ROW, True, HO_INTEREST_ROW)
        # Taxes Paid — studios + HO (studios are 0; HO has $1,188.60)
        pl.cell(row=PL_TAXES_ROW, column=col).value = studio_sum(cl, STUDIO_TAXES_ROW, True, HO_TAXES_ROW)
    print(f"    r{PL_RETAIL_ROW} Retail Sales        → SUM(studios r{STUDIO_RETAIL_ROW})")
    print(f"    r{PL_DEPRECIATION_ROW} Depreciation     → SUM(studios r{STUDIO_DEPRECIATION_ROW} + HO r{HO_DEPRECIATION_ROW})")
    print(f"    r{PL_INTEREST_ROW} Interest Expense → SUM(studios r{STUDIO_INTEREST_ROW} + HO r{HO_INTEREST_ROW})")
    print(f"    r{PL_TAXES_ROW} Taxes Paid       → SUM(studios r{STUDIO_TAXES_ROW} + HO r{HO_TAXES_ROW})")


def link_cash_flow_forecast(wb):
    """Step 5: rewire Cash Flow Forecast forecast rows to consolidated P&L."""
    print("\n[5] Cash Flow Forecast forecast → consolidated P&L (accrual = cash)")
    cf = wb["Cash Flow Forecast"]
    for col in range(JUN_2026, DEC_2028 + 1):
        cl = get_column_letter(col)
        cf.cell(row=CF_PROPERTY_COSTS_ROW, column=col).value = f"='P&L'!{cl}{PL_PROPERTY_COSTS_ROW}"
        cf.cell(row=CF_STAFF_COSTS_ROW, column=col).value = f"='P&L'!{cl}{PL_PAYROLL_ROW}"
        cf.cell(row=CF_UTILITIES_ROW, column=col).value = f"='P&L'!{cl}{PL_UTILITIES_ROW}"
        cf.cell(row=CF_MARKETING_ROW, column=col).value = f"='P&L'!{cl}{PL_MARKETING_ROW}"
        cf.cell(row=CF_ADMIN_ROW, column=col).value = f"='P&L'!{cl}{PL_SOFTWARE_ROW}"
        cf.cell(row=CF_PROFESSIONAL_FEES_ROW, column=col).value = f"='P&L'!{cl}{PL_PROFESSIONAL_FEES_ROW}"
        # Travel & Meals: no consolidated row — pull SUM directly from studio + HO subtotals
        cf.cell(row=CF_TRAVEL_MEALS_ROW, column=col).value = studio_sum(cl, STUDIO_TRAVEL_MEALS_TOTAL_ROW, True, HO_TRAVEL_MEALS_TOTAL_ROW)
        cf.cell(row=CF_MERCHANT_COGS_ROW, column=col).value = f"='P&L'!{cl}{PL_COGS_ROW}"
    print(f"    r{CF_PROPERTY_COSTS_ROW} Property Costs        → P&L r{PL_PROPERTY_COSTS_ROW}")
    print(f"    r{CF_STAFF_COSTS_ROW} Staff Costs           → P&L r{PL_PAYROLL_ROW} (Payroll)")
    print(f"    r{CF_UTILITIES_ROW} Utilities             → P&L r{PL_UTILITIES_ROW}")
    print(f"    r{CF_MARKETING_ROW} Marketing & Promotion → P&L r{PL_MARKETING_ROW}")
    print(f"    r{CF_ADMIN_ROW} Administrative & G&A  → P&L r{PL_SOFTWARE_ROW} (Software & Admin)")
    print(f"    r{CF_PROFESSIONAL_FEES_ROW} Professional Fees     → P&L r{PL_PROFESSIONAL_FEES_ROW}")
    print(f"    r{CF_TRAVEL_MEALS_ROW} Travel & Meals        → SUM(studios r50 + HO r30) [no consol row]")
    print(f"    r{CF_MERCHANT_COGS_ROW} Merchant Fees & COGS  → P&L r{PL_COGS_ROW}")
    print(f"    r19 Taxes               → P&L r25+r26 (already linked in v2)")
    print(f"    r18 Studio Start Up Costs → unchanged (=0)")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v2 file missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v2 → {OUT.name}")

    wb = load_workbook(OUT)
    print(f"Opened workbook: {len(wb.sheetnames)} sheets")

    move_interest_formula_to_ho(wb)
    update_ho_operating_tax(wb)
    zero_studio_taxes_forecast(wb)
    link_consolidated_pl(wb)
    link_cash_flow_forecast(wb)

    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
