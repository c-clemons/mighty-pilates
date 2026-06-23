"""
Apply Cat's authoritative May 2026 cash sales per studio.

Cat provided the May 2026 cash sales totals per studio (her image-table dated
2026-06-23, total = $799,443). These supersede our prior forecasted values
in client_sales_forecast for May 2026.

Updates:
  1. dashboard/data/committed_actuals.json
       - client_sales_forecast[CODE]['2026-05']                = Cat's value
       - client_sales_forecast_consolidated['2026-05']         = $799,443
       - monthly_sales['2026-05']                              = $799,443
  2. Excel Sales Forecast tab (rows 6-17, col H = May 2026 (A))
  3. Excel Cash Flow Forecast tab (row 7 H = Total Cash Sales) — replace
     stale hardcoded value with formula referencing Sales Forecast row 18
  4. Save Excel to BOTH:
       - /Users/chandlerclemons/Desktop/Empirica Financial Modeling/Mighty Pilates/
       - /Users/chandlerclemons/Desktop/Mighty Pilates/
  5. Snapshot Excel into repo (snapshots/excel/)
"""
from __future__ import annotations
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Cat's May 2026 cash sales per studio (image 2026-06-23, "Total Sales" column)
CAT_MAY_2026 = {
    "BK":  83273,
    "CC":  43327,
    "DN":  15337,
    "LF":  71444,
    "MR":  74382,
    "OP":  63770,
    "PH":  93511,
    "RH":  66857,
    "SB":  20133,
    "SM": 206632,
    "WP":     26,
    "WW":  60750,
}
CAT_TOTAL = sum(CAT_MAY_2026.values())  # 799,442 (off by $1 from Cat's reported $799,443 rounding)
TARGET_TOTAL = 799443                    # Use Cat's exact reported total

MONTH_DASH = "2026-05"
EXCEL_COL_INDEX = 8  # H = May 2026 (A) on Sales Forecast

REPO_ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_JSON = REPO_ROOT / "dashboard" / "data" / "committed_actuals.json"

EXCEL_LIVE_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Mighty Pilates/"
    "Mighty Pilates Financial Model.xlsx"
)
EXCEL_SECONDARY_PATH = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Pilates Financial Model.xlsx")
EXCEL_REPO_SNAPSHOT = REPO_ROOT / "snapshots" / "excel" / "Mighty_Pilates_Financial_Model_May2026.xlsx"


def update_dashboard_json():
    print(f"\n[1/4] Updating {DASHBOARD_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(DASHBOARD_JSON.read_text())

    csf = data.setdefault("client_sales_forecast", {})
    csfc = data.setdefault("client_sales_forecast_consolidated", {})
    ms = data.setdefault("monthly_sales", {})

    for code, amt in CAT_MAY_2026.items():
        bucket = csf.setdefault(code, {})
        old = bucket.get(MONTH_DASH, 0)
        bucket[MONTH_DASH] = float(amt)
        print(f"   client_sales_forecast[{code}][{MONTH_DASH}]: ${old:>12,.2f} → ${amt:>12,.2f}")

    old_total = csfc.get(MONTH_DASH, 0)
    csfc[MONTH_DASH] = float(TARGET_TOTAL)
    print(f"   client_sales_forecast_consolidated[{MONTH_DASH}]: ${old_total:>12,.2f} → ${TARGET_TOTAL:>12,.2f}")

    old_ms = ms.get(MONTH_DASH, 0)
    ms[MONTH_DASH] = float(TARGET_TOTAL)
    print(f"   monthly_sales[{MONTH_DASH}]:                     ${old_ms:>12,.2f} → ${TARGET_TOTAL:>12,.2f}")

    data["_last_updated"] = datetime.now().isoformat()
    DASHBOARD_JSON.write_text(json.dumps(data, indent=2))
    print(f"   ✓ Saved")


def update_excel():
    print(f"\n[2/4] Updating Excel: {EXCEL_LIVE_PATH.name}")
    wb = load_workbook(EXCEL_LIVE_PATH)

    # Sales Forecast tab — rows 6-17, col H (= 8) is May 2026
    sf = wb["Sales Forecast"]
    code_to_row = {}
    for r in range(6, 18):
        code = sf.cell(row=r, column=2).value  # col B is Code
        if code:
            code_to_row[str(code).strip()] = r

    for code, amt in CAT_MAY_2026.items():
        if code not in code_to_row:
            print(f"   WARNING: studio code {code!r} not found in Sales Forecast tab")
            continue
        r = code_to_row[code]
        cell = sf.cell(row=r, column=EXCEL_COL_INDEX)
        old = cell.value
        cell.value = float(amt)
        old_str = f"${float(old):>12,.2f}" if isinstance(old, (int, float)) else f"{old}"
        print(f"   SalesForecast!H{r}  ({code}): {old_str:>16} → ${amt:>12,.2f}")

    # Row 18 has SUM formula — recomputes automatically when Excel opens. Confirm formula intact.
    sum_cell = sf.cell(row=18, column=EXCEL_COL_INDEX)
    print(f"   SalesForecast!H18  (TOTAL formula): {sum_cell.value}")

    # Cash Flow Forecast tab — row 7 H was hard-coded to a stale value. Replace
    # with formula reference to Sales Forecast row 18 (consistent with adjacent cells).
    cf = wb["Cash Flow Forecast"]
    cf_cell = cf.cell(row=7, column=EXCEL_COL_INDEX)
    old_cf = cf_cell.value
    cf_cell.value = "='Sales Forecast'!H18"
    print(f"\n   CashFlowForecast!H7 Total Cash Sales: {old_cf!r} → ='Sales Forecast'!H18")

    # Update header text to reflect May is now actuals (was "May 2026 (A)" already, leave it)
    # And the explanation row
    explain_cell = sf.cell(row=3, column=2)
    if "Jan-Apr 2026" in str(explain_cell.value or ""):
        explain_cell.value = (
            "Jan-May 2026 actuals (locked, Cat-authoritative). "
            "Jun 2026 onwards forecast (editable)."
        )
        print(f"   SalesForecast!B3 header updated → 'Jan-May 2026 actuals (locked, Cat-authoritative)...'")

    return wb


def save_excel_everywhere(wb):
    print(f"\n[3/4] Saving Excel to 3 locations")
    paths = [EXCEL_LIVE_PATH, EXCEL_SECONDARY_PATH, EXCEL_REPO_SNAPSHOT]
    for p in paths:
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        print(f"   ✓ {p}")


def main():
    print("=" * 78)
    print(f"Apply Cat's May 2026 cash sales (total ${TARGET_TOTAL:,})")
    print("=" * 78)

    update_dashboard_json()
    wb = update_excel()
    save_excel_everywhere(wb)

    print(f"\n[4/4] Done.")
    print(f"      Per-studio sum of provided values: ${CAT_TOTAL:,}")
    print(f"      Cat's reported total:              ${TARGET_TOTAL:,}")
    print(f"      Rounding delta (TOTAL - sum):      ${TARGET_TOTAL - CAT_TOTAL}")
    print(f"      → Excel total = SUM(per-studio) = ${CAT_TOTAL:,}")
    print(f"      → Dashboard consolidated explicitly set to ${TARGET_TOTAL:,}")


if __name__ == "__main__":
    main()
