"""
v4 of the analyst's lender model.

v3 → v4 changes per Cat's 2026-06-30 follow-up:

1. Cash Flow Forecast actuals (Jan-May 2026) for operating expenses don't tie
   to the consolidated P&L equivalents (e.g., CF Merchant Fees & COGS Jan 2026
   = $65,117 hardcoded vs P&L = $38,010 actual). Per Cat's "accrual = cash"
   assumption, these should equal.
   FIX: Extend the v3 CF→P&L linking backward to cover Jan-May 2026 cols
   (P-T) so actuals tie. Cols D-O (2025) left None (analyst's original).

2. Property Taxes: keep on consolidated P&L r26 as hardcoded (no clean per-
   studio source) but RESET the forecast value to 2025 monthly average.
     2025 Property Taxes total: $50,138.23 / 12 = $4,178.19/mo
     (was: $3,673/mo trailing-3-mo avg from analyst)
   Add a note to the Read Me tab explaining this.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v3.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v4.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v4.xlsx"

# Column extent: Jan 2026 (col P=16) through May 2026 (col T=20)
JAN_2026 = 16
MAY_2026 = 20
JUN_2026 = 21
DEC_2028 = 39

STUDIO_TABS = ['BK P&L', 'CC P&L', 'DN P&L', 'LF P&L', 'MR P&L', 'OP P&L',
               'PH P&L', 'RH P&L', 'SB P&L', 'SM P&L', 'WP P&L', 'WW P&L']

# Same CF → P&L mapping as v3, but applied to actuals cols too
CF_TO_PL_MAP = {
    10: 19,   # Property Costs        → P&L Property Costs
    11: 15,   # Staff Costs           → P&L Payroll
    12: 18,   # Utilities             → P&L Utilities
    13: 14,   # Marketing & Promotion → P&L Marketing
    14: 16,   # Administrative & G&A  → P&L Software & Admin
    15: 17,   # Professional Fees     → P&L Professional Fees
    17: 12,   # Merchant Fees & COGS  → P&L COGS & Merchant Fees
}
# r16 Travel & Meals — special, SUM(studios r50 + HO r30)
# r19 Taxes — special, P&L r25 + r26

# Property Tax — use 2025 monthly average
PROP_TAX_2025_TOTAL = 50138.23
PROP_TAX_NEW_MONTHLY = round(PROP_TAX_2025_TOTAL / 12, 2)  # $4,178.19


def link_cf_actuals_to_pl(wb):
    """Extend CF→P&L linkage backward to cover Jan-May 2026 actuals (cols P-T)."""
    print("\n[1] CF actuals (Jan-May 2026) → consolidated P&L")
    cf = wb["Cash Flow Forecast"]
    for col in range(JAN_2026, MAY_2026 + 1):
        cl = get_column_letter(col)
        for cf_row, pl_row in CF_TO_PL_MAP.items():
            cf.cell(row=cf_row, column=col).value = f"='P&L'!{cl}{pl_row}"
        # Travel & Meals — SUM(studios r50 + HO r30)
        parts = [f"'{tab}'!{cl}50" for tab in STUDIO_TABS] + [f"'HO P&L'!{cl}30"]
        cf.cell(row=16, column=col).value = "=" + "+".join(parts)
        # Taxes (r19) — P&L r25 + r26
        cf.cell(row=19, column=col).value = f"='P&L'!{cl}25+'P&L'!{cl}26"
    print(f"    Updated cols P-T (Jan-May 2026) for CF rows 10-17 and 19")
    print(f"    Now ties cleanly to consolidated P&L for all months Jan 2026+")


def update_property_tax_forecast(wb):
    """P&L r26 Property Taxes forecast → 2025 monthly average ($4,178.19/mo)."""
    print(f"\n[2] P&L r26 Property Taxes forecast → ${PROP_TAX_NEW_MONTHLY}/mo (2025 monthly avg)")
    pl = wb["P&L"]
    for col in range(JUN_2026, DEC_2028 + 1):
        pl.cell(row=26, column=col).value = PROP_TAX_NEW_MONTHLY
    print(f"    2025 Property Tax total: ${PROP_TAX_2025_TOTAL:,.2f} / 12 = ${PROP_TAX_NEW_MONTHLY}/mo")
    print(f"    Cols U-AM (Jun 2026 - Dec 2028) updated")


def add_property_tax_note(wb):
    """Append a property-tax note to the Read Me tab."""
    print("\n[3] Read Me note: Property Tax assumption")
    rm = wb["Read Me"]
    # Find next empty row after the v2 'Cash Sales vs Accrual' note (rows 45-50)
    next_row = 52
    rm.cell(row=next_row, column=2, value="Property Tax — forecast assumption").font = Font(bold=True, size=11)
    notes = [
        ("Why hardcoded", "Property Tax on consolidated P&L row 26 is kept as a hardcoded forecast value rather than linked to individual P&L tabs because the QBO source booked property tax at corporate (Norbrook) level, not per-studio. Studio P&L row 68 ('903000 Property taxes') is $0 in actuals."),
        ("Forecast basis", f"Forecast value ${PROP_TAX_NEW_MONTHLY}/mo = 2025 actual annual Property Taxes (${PROP_TAX_2025_TOTAL:,.2f}) / 12. Conservative assumption that 2026 onwards will run at the same rate."),
        ("If you want it linked", "Would require either (a) inserting a Property Tax row on HO P&L (structural change with downstream formula updates), or (b) allocating across studios by some basis (revenue weight, square footage, etc.). Either is doable — let me know."),
    ]
    for i, (label, body) in enumerate(notes, start=1):
        rm.cell(row=next_row + i, column=2, value=label).font = Font(bold=True)
        rm.cell(row=next_row + i, column=3, value=body).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"    Added rows {next_row}-{next_row + len(notes)}")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v3 file missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v3 → {OUT.name}")

    wb = load_workbook(OUT)
    link_cf_actuals_to_pl(wb)
    update_property_tax_forecast(wb)
    add_property_tax_note(wb)

    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
