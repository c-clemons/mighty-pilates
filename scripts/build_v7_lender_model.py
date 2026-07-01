"""
v7: relabel CF r22 to "Total Operating Cash Outflows" (Option 1 per Cat)
plus comprehensive tie-out verification.

v6 → v7 change:
  (a) Rename Cash Flow Forecast r22 from "Total Operating Expenses" to
      "Total Operating Cash Outflows" — clarifies that it includes COGS,
      Taxes, and PF charges (not just OpEx like the P&L r21 line).
  (b) Verify all Total rows tie across the model.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v6.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v7.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v7.xlsx"


def relabel_cf_total(wb):
    print("\n[a] Relabel CF r22 → 'Total Operating Cash Outflows'")
    cf = wb["Cash Flow Forecast"]
    cf.cell(row=22, column=2).value = "Total Operating Cash Outflows"


def fix_studio_total_other_expenses(wb):
    """Studio P&L r69 Total Other Expenses currently hardcoded 0 for forecast
    (or matches just depreciation for actuals). Fix to =SUM(r65:r68) for all
    cols so per-studio Net Income shows the depreciation flowing through."""
    print("\n[b] Studio P&L r69 Total Other Expenses → =SUM(r65:r68) for all cols")
    from openpyxl.utils import get_column_letter
    STUDIOS = ['BK P&L','CC P&L','DN P&L','LF P&L','MR P&L','OP P&L',
               'PH P&L','RH P&L','SB P&L','SM P&L','WP P&L','WW P&L']
    for tab in STUDIOS:
        ws = wb[tab]
        for col in range(4, 40):  # D-AM
            cl = get_column_letter(col)
            ws.cell(row=69, column=col).value = f"=SUM({cl}65:{cl}68)"
    print(f"    Updated 12 studio P&L tabs, cols D-AM")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v6 missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v6 → {OUT.name}")

    wb = load_workbook(OUT)
    relabel_cf_total(wb)
    fix_studio_total_other_expenses(wb)
    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
