"""
Apply Cat's authoritative July 2026 cash sales per studio.

Source: Cat's July 2026 studio summary (2026-08-04), "Total Sales" column
(= MindBody sales + ClassPass + Wellhub + Retail, i.e. total cash sales).
Per-studio sum: $837,634 (Cat's reported grand total: $837,632, $2 rounding).

Reconciliation (scripts/jul2026_sales_reconcile.py, 2026-08-04):
  - MB-side (MART_SALES_DETAILS NET_PAYMENTAMT_LOCAL) ties to Cat within -0.38%
    consolidated; 4 studios flagged >$1K (SM +3.1K, PH -1.5K, RH -1.1K, LF -1.1K)
    for Cat follow-up.
  - ClassPass (RESERVATIONS) only loaded through 2026-07-26 at apply time, so the
    live CP total read ~$30K light vs Cat. Load lag, not a real discrepancy.
  Chandler directed applying Cat's authoritative Totals now regardless of the CP lag.

Updates:
  1. dashboard/data/baseline.json           (git-tracked persistent state)
  2. dashboard/data/user_overrides.json     (Cash Flow page reads this)
  3. dashboard/data/committed_actuals.json  (client_sales_forecast tree)
  4. Excel Sales Forecast tab, col J = Jul 2026
  5. Excel Cash Flow Forecast row 7 → formula reference to Sales Forecast
"""
from __future__ import annotations
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Cat's July 2026 "Total Sales" per studio (includes ClassPass).
CAT_JUL_2026 = {
    "BK":  99551,
    "CC":  42903,
    "DN":  16795,
    "LF":  86878,
    "MR": 109856,
    "OP":  58162,
    "PH": 117515,
    "RH":  75936,
    "SB":  18714,
    "SM": 163688,
    "WP":     99,
    "WW":  47537,
}
CAT_TOTAL = sum(CAT_JUL_2026.values())    # 837,634
TARGET_TOTAL = 837632                     # Cat's reported grand-total row

MONTH_DASH = "2026-07"
EXCEL_COL_INDEX = 10  # J = Jul 2026

REPO_ROOT = Path(__file__).resolve().parents[1]
BASELINE_JSON = REPO_ROOT / "dashboard" / "data" / "baseline.json"
USER_OVERRIDES_JSON = REPO_ROOT / "dashboard" / "data" / "user_overrides.json"
DASHBOARD_JSON = REPO_ROOT / "dashboard" / "data" / "committed_actuals.json"

EXCEL_LIVE_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Mighty Pilates/"
    "Mighty Pilates Financial Model.xlsx"
)
EXCEL_SECONDARY_PATH = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Pilates Financial Model.xlsx")
EXCEL_REPO_SNAPSHOT = REPO_ROOT / "snapshots" / "excel" / "Mighty_Pilates_Financial_Model_Jul2026.xlsx"


def update_baseline():
    print(f"\n[A0] Updating {BASELINE_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(BASELINE_JSON.read_text())
    sf = data.setdefault("sales_forecast", {})
    for code, amt in CAT_JUL_2026.items():
        bucket = sf.setdefault(code, {})
        old = bucket.get(MONTH_DASH, 0)
        bucket[MONTH_DASH] = float(amt)
        print(f"   baseline[{code}][{MONTH_DASH}]: ${old:>12,.2f} → ${amt:>12,.2f}")
    BASELINE_JSON.write_text(json.dumps(data, indent=2))


def update_user_overrides():
    print(f"\n[A] Updating {USER_OVERRIDES_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(USER_OVERRIDES_JSON.read_text())
    sf = data.setdefault("sales_forecast", {})
    for code, amt in CAT_JUL_2026.items():
        bucket = sf.setdefault(code, {})
        old = bucket.get(MONTH_DASH, 0)
        bucket[MONTH_DASH] = float(amt)
        print(f"   user_overrides[{code}][{MONTH_DASH}]: ${old:>12,.2f} → ${amt:>12,.2f}")
    data["_last_updated"] = datetime.now().isoformat()
    USER_OVERRIDES_JSON.write_text(json.dumps(data, indent=2))


def update_dashboard_json():
    print(f"\n[1/4] Updating {DASHBOARD_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(DASHBOARD_JSON.read_text())
    csf = data.setdefault("client_sales_forecast", {})
    csfc = data.setdefault("client_sales_forecast_consolidated", {})
    ms = data.setdefault("monthly_sales", {})
    for code, amt in CAT_JUL_2026.items():
        bucket = csf.setdefault(code, {})
        bucket[MONTH_DASH] = float(amt)
    csfc[MONTH_DASH] = float(TARGET_TOTAL)
    ms[MONTH_DASH] = float(TARGET_TOTAL)
    print(f"   client_sales_forecast_consolidated[{MONTH_DASH}] = ${TARGET_TOTAL:,}")
    print(f"   monthly_sales[{MONTH_DASH}]                     = ${TARGET_TOTAL:,}")
    data["_last_updated"] = datetime.now().isoformat()
    DASHBOARD_JSON.write_text(json.dumps(data, indent=2))


def update_excel():
    print(f"\n[2/4] Updating Excel: {EXCEL_LIVE_PATH.name}")
    if not EXCEL_LIVE_PATH.exists():
        raise FileNotFoundError(f"Live Excel not found: {EXCEL_LIVE_PATH}")
    wb = load_workbook(EXCEL_LIVE_PATH)

    sf = wb["Sales Forecast"]
    code_to_row = {}
    for r in range(6, 18):
        code = sf.cell(row=r, column=2).value
        if code:
            code_to_row[str(code).strip()] = r

    for code, amt in CAT_JUL_2026.items():
        if code not in code_to_row:
            print(f"   WARNING: studio code {code!r} not found in Sales Forecast tab")
            continue
        r = code_to_row[code]
        cell = sf.cell(row=r, column=EXCEL_COL_INDEX)
        old = cell.value
        cell.value = float(amt)
        old_str = f"${float(old):>12,.2f}" if isinstance(old, (int, float)) else f"{old}"
        print(f"   SalesForecast!J{r}  ({code}): {old_str:>16} → ${amt:>12,.2f}")

    sum_cell = sf.cell(row=18, column=EXCEL_COL_INDEX)
    print(f"   SalesForecast!J18 (TOTAL formula): {sum_cell.value}")

    # Mark the Jul column header as Actual (May uses the '(A)' convention).
    hdr = sf.cell(row=5, column=EXCEL_COL_INDEX)
    if hdr.value and "(A)" not in str(hdr.value):
        old_hdr = hdr.value
        hdr.value = f"{str(hdr.value).strip()} (A)"
        print(f"   SalesForecast!J5 header: {old_hdr!r} → {hdr.value!r}")

    # Cash Flow Forecast row 7, col J = Jul 2026 — formula reference (idempotent).
    cf = wb["Cash Flow Forecast"]
    cf_cell = cf.cell(row=7, column=EXCEL_COL_INDEX)
    old_cf = cf_cell.value
    cf_cell.value = "='Sales Forecast'!J18"
    print(f"\n   CashFlowForecast!J7 Total Cash Sales: {old_cf!r} → ='Sales Forecast'!J18")

    # Update the B3 header note to reflect July is now actuals (Aug+ forecast).
    explain_cell = sf.cell(row=3, column=2)
    val = str(explain_cell.value or "")
    if "Jul 2026 actuals" not in val:
        explain_cell.value = (
            "Jan-Jul 2026 actuals (locked, Cat-authoritative). "
            "Aug 2026 onwards forecast (editable)."
        )
        print(f"   SalesForecast!B3 header → 'Jan-Jul 2026 actuals (locked)...'")

    return wb


def save_excel_everywhere(wb):
    print(f"\n[3/4] Saving Excel to 3 locations")
    for p in [EXCEL_LIVE_PATH, EXCEL_SECONDARY_PATH, EXCEL_REPO_SNAPSHOT]:
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        print(f"   ✓ {p}")


def main():
    print("=" * 78)
    print(f"Apply Cat's July 2026 cash sales (total ${TARGET_TOTAL:,})")
    print("=" * 78)
    update_baseline()
    update_user_overrides()
    update_dashboard_json()
    wb = update_excel()
    save_excel_everywhere(wb)
    print(f"\n[4/4] Done.")
    print(f"      Per-studio sum: ${CAT_TOTAL:,}")
    print(f"      Reported total: ${TARGET_TOTAL:,}")
    print(f"      Rounding delta: ${TARGET_TOTAL - CAT_TOTAL}")


if __name__ == "__main__":
    main()
