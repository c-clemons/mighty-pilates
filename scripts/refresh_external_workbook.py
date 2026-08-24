"""
Stage 3 — push a new actuals month into the external financial workbook.

This replaces the retired internal-model path
(`financial-modeling/models/mighty/refresh_from_streamlit.py`), which had not
been run since the April 2026 actuals and lives in a repo that has since moved
on to another client. The external workbook is now the model of record.

WHAT IT TOUCHES
---------------
1. `QBO Actuals`  — appends the new month's column to all three blocks
                    (P&L, Balance Sheet, Cash Flow), read straight from
                    committed_actuals.json.
2. `P&L (Existing Locations)` — converts the new month's column from forecast
                    formulas to actuals by translating the prior (actual)
                    column one to the right. Because the prior column reads
                    `'QBO Actuals'!<col><row>`, the translation advances that
                    reference to the new month automatically.
3. `<XX> P&L` studio tabs — writes hardcoded values for the new month, matching
                    the convention already used in those tabs' actuals columns.
                    Subtotal rows keep the workbook's own formulas (translated).
4. `Assumptions`  — bumps "Last Actuals Month".
5. `P&L (Existing Locations)` header note and `Cover` refresh date.

SAFETY
------
`--validate` recomputes an ALREADY-POPULATED actuals month from
committed_actuals.json and diffs it against what the workbook holds. Run it
before every real refresh. If the studio row map in workbook_studio_map.py has
drifted from Crew's chart of accounts, the validator catches it there instead of
silently writing wrong numbers.

USAGE
-----
    python scripts/refresh_external_workbook.py --validate --month "Jun 2026"
    python scripts/refresh_external_workbook.py --month "Jul 2026" \
        --source "<path to prior workbook>" --out "<path to new workbook>"
"""
from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter, column_index_from_string

from workbook_studio_map import (normalize, studio_row_value, ho_row_value,
                                 manual_adjustment)

REPO_ROOT = Path(__file__).resolve().parents[1]
DASHBOARD_JSON = REPO_ROOT / "dashboard" / "data" / "committed_actuals.json"

QBO_SHEET = "QBO Actuals"
CONSOL_SHEET = "P&L (Existing Locations)"
STUDIO_TABS = [
    "BK P&L", "CC P&L", "DN P&L", "LF P&L", "MR P&L", "OP P&L",
    "PH P&L", "RH P&L", "SB P&L", "SM P&L", "WP P&L", "WW P&L", "HO P&L",
]

# QBO Actuals is three stacked blocks. Each has its own header row; the rows
# beneath carry account labels in column B that match committed_actuals keys.
# Section titles in column B that open each block. Boundaries are resolved at
# runtime because inserting a new account row shifts everything below it.
QBO_BLOCK_TITLES = [("pl", "P&L"), ("bs", "Balance Sheet"), ("scf", "Cash Flow Statement")]


def locate_blocks(ws):
    """-> [(section, header_row, end_row)], resolved from the sheet itself."""
    starts = []
    for section, title in QBO_BLOCK_TITLES:
        for r in range(1, ws.max_row + 1):
            if normalize(ws.cell(r, 2).value) == title:
                starts.append((section, r + 1))   # header (month) row
                break
    out = []
    for i, (section, header_row) in enumerate(starts):
        end = starts[i + 1][1] - 2 if i + 1 < len(starts) else ws.max_row
        out.append((section, header_row, end))
    return out

TOLERANCE = 0.02

# The workbook keeps these bare subtotal-label rows at 0 (QBO native); the value
# lives on the sibling "Total ..." row. dashboard_update.py injects the total
# into the bare label, so we must NOT propagate that into the workbook.
QBO_SKIP_LABELS = {"401000 Sessions", "403000 Breakage Revenue"}

# Accounts Crew added that have no row in the workbook's P&L block. Written to
# free rows at the end of the block (nothing references them, so no formula
# breaks and no row numbers shift).
QBO_EXTRA_ROWS = {"401007 Off-Site": 92}

# Accounts Crew has added that the workbook has no row for. Without a row, the
# account's value still reaches the "Total ..." rows (those are written straight
# from the JSON) but no visible component row carries it — so the total stops
# equalling the sum of the rows above it. Each entry inserts a row in the right
# place. Add to this list whenever --audit reports a missing account.
#   (section, account label as it appears in committed_actuals, insert AFTER this row's label)
NEW_ACCOUNT_ROWS = [
    ("bs", "131120 Prepaid Property Tax", "131100 Prepaid expenses"),
    ("bs", "155009 Leasehold Improvements - Presidio Heights",
           "155008 Leasehold Improvements - Ocean Park"),
    ("bs", "242250 Khary Loan #NA", "242200 Specialty Capital Loan"),
    ("scf", "131120 Prepaid expenses:Prepaid Property Tax", "131100 Prepaid expenses"),
    ("scf", "155009 Fixed Assets:Leasehold Improvements:Leasehold Improvements - Presidio Heights",
            "151000 Fixed Assets:Furniture & Fixtures"),
]


def _find_row(ws, label: str, lo: int = 1, hi: int | None = None) -> int | None:
    """Row whose column-B label matches, searched within [lo, hi].

    The range matters: the same label (e.g. "131100 Prepaid expenses") appears
    in more than one block, so an unscoped search lands in the wrong section.
    """
    target = normalize(label)
    for r in range(lo, (hi or ws.max_row) + 1):
        if normalize(ws.cell(r, 2).value) == target:
            return r
    return None


def _shift_qbo_refs(wb, at_row: int) -> int:
    """Bump every 'QBO Actuals'!<col><row> reference at/below at_row by one.

    openpyxl's insert_rows moves cells but leaves formulas untouched, so any
    sheet pointing into QBO Actuals would silently read the wrong row after an
    insert. QBO Actuals itself holds no formulas, so only cross-sheet refs
    matter.
    """
    pattern = re.compile(r"('QBO Actuals'!\$?)([A-Z]{1,3})(\$?)(\d+)")

    def bump(m):
        row = int(m.group(4))
        return f"{m.group(1)}{m.group(2)}{m.group(3)}{row + 1 if row >= at_row else row}"

    fixed = 0
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and "QBO Actuals" in v:
                    new = pattern.sub(bump, v)
                    if new != v:
                        cell.value = new
                        fixed += 1
    return fixed


def insert_missing_accounts(wb, data, blocks) -> set[str]:
    """Give newly-added accountant accounts a row. Returns the labels added."""
    ws = wb[QBO_SHEET]
    bounds = {sec: (hdr, end) for sec, hdr, end in locate_blocks(ws)}
    planned = []
    for section, label, after in NEW_ACCOUNT_ROWS:
        lo, hi = bounds[section]
        if _find_row(ws, label, lo, hi) is not None:
            continue
        anchor = _find_row(ws, after, lo, hi)
        if anchor is None:
            print(f"  !! cannot place '{label}' in {section}: "
                  f"anchor '{after}' not found in that block")
            continue
        planned.append((anchor + 1, section, label))

    added = set()
    # Descending, so an insert never invalidates a lower-numbered target.
    for at, section, label in sorted(planned, reverse=True):
        ws.insert_rows(at)
        ws.cell(at, 2).value = label
        ws.cell(at, 2)._style = ws.cell(at + 1, 2)._style
        n = _shift_qbo_refs(wb, at)
        added.add(normalize(label))
        print(f"  {QBO_SHEET}/{section}: inserted r{at} '{label[:44]}' "
              f"({n} cross-sheet refs repaired)")
    return added


def _lookup(section_data: dict, label: str):
    """Resolve a workbook row label against a committed_actuals section.

    Crew renamed the balance-sheet/cash-flow subtotal rows in the Jul 2026
    package ("Total for Assets" -> "TOTAL ASSETS", "Total for Bank Accounts" ->
    "Total Bank Accounts"), and some keys carry stray double spaces. Match
    forgivingly so a cosmetic rename on their side does not blank a row here.
    Returns None when there is genuinely no match.
    """
    norm = {normalize(k): v for k, v in section_data.items()}
    if label in norm:
        return norm[label]
    lower = {k.lower(): v for k, v in norm.items()}
    for cand in (label, label.replace("Total for ", "Total ")):
        if cand.lower() in lower:
            return lower[cand.lower()]
    return None


def studio_code(tab: str) -> str:
    return tab.replace(" P&L", "").strip()


def strip_marker(label: str) -> str:
    """'Jun 2026 (A)' -> 'Jun 2026'. Some tabs tag actual/forecast columns."""
    return label.split("(")[0].strip() if "(" in label else label


def find_month_col(ws, header_row: int, month: str) -> int | None:
    for c in range(2, ws.max_column + 2):
        if strip_marker(normalize(ws.cell(header_row, c).value)) == month:
            return c
    return None


def _month_cols(ws, header_row: int):
    """Yield (month_label, column_index) for every month column in a block."""
    for c in range(2, ws.max_column + 2):
        v = strip_marker(normalize(ws.cell(header_row, c).value))
        if v and v != "Account" and v.split()[-1].isdigit():
            yield v, c


def prior_month(month: str) -> str:
    mons = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    m, y = month.split()
    i = mons.index(m)
    return f"{mons[i - 1]} {y}" if i else f"Dec {int(y) - 1}"


# ---------------------------------------------------------------- validate ---

def validate(wb, data, month: str) -> int:
    """Diff a populated actuals month in the workbook against the JSON."""
    problems = 0
    print("=" * 74)
    print(f"VALIDATE — recomputing {month} from committed_actuals.json")
    print("=" * 74)

    # --- QBO Actuals blocks
    ws = wb[QBO_SHEET]
    for section, header_row, end_row in locate_blocks(ws):
        col = find_month_col(ws, header_row, month)
        if col is None:
            print(f"  {QBO_SHEET}/{section}: no '{month}' column — skipped")
            continue
        checked = mismatched = 0
        for r in range(header_row + 1, end_row + 1):
            label = normalize(ws.cell(r, 2).value)
            if not label or label == "Account":
                continue
            if label in QBO_SKIP_LABELS:
                continue
            found = _lookup(data[section].get(month, {}), label)
            if found is None:
                continue
            want = round(float(found or 0), 2)
            got = ws.cell(r, col).value
            got = round(float(got or 0), 2) if not isinstance(got, str) else None
            checked += 1
            if got is None or abs(got - want) > TOLERANCE:
                mismatched += 1
                problems += 1
                if mismatched <= 6:
                    print(f"    MISMATCH {section} r{r} {label[:44]:<46}"
                          f"workbook={got} json={want}")
        print(f"  {QBO_SHEET}/{section:<4} checked {checked:>4}  "
              f"mismatched {mismatched}")

    # --- reverse coverage: accountant rows the workbook has nowhere to put.
    # Without this the totals (written straight from JSON) stay correct while
    # the component rows above them quietly fail to add up. This is how Crew's
    # Jul-2026 additions (131120 Prepaid Property Tax, 155009 Leasehold - PH,
    # 242250 Khary Loan #NA) were caught. Fix by adding to NEW_ACCOUNT_ROWS.
    for section, header_row, end_row in locate_blocks(ws):
        rows = {normalize(ws.cell(r, 2).value)
                for r in range(header_row + 1, end_row + 1)}
        rows |= {x.replace("Total for ", "Total ") for x in rows}
        lower = {x.lower() for x in rows}
        for label, value in data[section].get(month, {}).items():
            nk = normalize(label)
            if nk.lower() in lower or nk.replace("Total ", "Total for ").lower() in lower:
                continue
            if not value:
                continue
            if nk.startswith(("Total", "TOTAL")):
                # A subtotal the workbook does not present. Harmless as long as
                # the enclosing total still sums its components, which the
                # arithmetic audit covers separately.
                continue
            print(f"    NO WORKBOOK ROW  {section} {label[:50]:<52}{value:>14,.2f}")
            problems += 1

    # --- studio tabs
    for tab in STUDIO_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        code = studio_code(tab)
        month_data = data["studios"].get(code, {}).get("data", {}).get(month)
        if not month_data:
            print(f"  {tab:<10} no JSON data for {month} — skipped")
            continue
        col = find_month_col(ws, 5, month)
        if col is None:
            print(f"  {tab:<10} no '{month}' column — skipped")
            continue
        checked = mismatched = 0
        unmapped = []
        for r in range(6, ws.max_row + 1):
            label = ws.cell(r, 2).value
            if label is None:
                continue
            if tab == "HO P&L":
                want = ho_row_value(month_data, r)
            else:
                try:
                    want = studio_row_value(month_data, label)
                except KeyError as exc:
                    unmapped.append(f"r{r} {exc.args[0]}")
                    problems += 1
                    continue
            if want is None:
                continue
            want = round(want + manual_adjustment(tab, r), 2)
            got = ws.cell(r, col).value
            if isinstance(got, str):
                continue  # a formula where we expect a value; report separately
            got = round(float(got or 0), 2)
            checked += 1
            if abs(got - want) > TOLERANCE:
                mismatched += 1
                problems += 1
                if mismatched <= 5:
                    print(f"    MISMATCH {tab} r{r} {normalize(label)[:40]:<42}"
                          f"workbook={got:>12,.2f} json={want:>12,.2f}")
        flag = "" if not (mismatched or unmapped) else "  <-- CHECK"
        print(f"  {tab:<10} checked {checked:>3}  mismatched {mismatched:>3}"
              f"  unmapped {len(unmapped)}{flag}")
        for u in unmapped[:6]:
            print(f"      UNMAPPED {u}")

    print("=" * 74)
    print("PASS — map reproduces the workbook" if problems == 0
          else f"FAIL — {problems} problem(s); fix workbook_studio_map.py")
    print("=" * 74)
    return problems


# ----------------------------------------------------------------- refresh ---

def refresh(wb, data, month: str) -> None:
    prev = prior_month(month)
    print("=" * 74)
    print(f"REFRESH — adding {month} (prior actuals month: {prev})")
    print("=" * 74)

    # 1. QBO Actuals: append the month column to each block.
    ws = wb[QBO_SHEET]
    added = insert_missing_accounts(wb, data, None)

    for section, header_row, end_row in locate_blocks(ws):
        prev_col = find_month_col(ws, header_row, prev)
        if prev_col is None:
            raise SystemExit(f"{QBO_SHEET}/{section}: prior month '{prev}' not found")
        col = find_month_col(ws, header_row, month) or prev_col + 1
        ws.cell(header_row, col).value = month
        ws.cell(header_row, col)._style = ws.cell(header_row, prev_col)._style
        written = missing = 0
        unmatched: list[str] = []
        for r in range(header_row + 1, end_row + 1):
            label = normalize(ws.cell(r, 2).value)
            if not label or label == "Account":
                continue
            if label in QBO_SKIP_LABELS:
                ws.cell(r, col).value = 0
                ws.cell(r, col)._style = ws.cell(r, prev_col)._style
                continue
            found = _lookup(data[section].get(month, {}), label)
            if found is not None:
                ws.cell(r, col).value = round(float(found or 0), 2)
                written += 1
                if label in added:
                    # brand-new row: fill its prior months too, so the column
                    # above it is not blank
                    for m2, c2 in _month_cols(ws, header_row):
                        v2 = _lookup(data[section].get(m2, {}), label)
                        ws.cell(r, c2).value = round(float(v2 or 0), 2)
            else:
                unmatched.append(f"r{r} {label}")
                missing += 1
            ws.cell(r, col)._style = ws.cell(r, prev_col)._style
        if section == "pl":
            for label, row in QBO_EXTRA_ROWS.items():
                ws.cell(row, 2).value = label
                ws.cell(row, col).value = round(
                    float(data["pl"].get(month, {}).get(label, 0) or 0), 2)
                # backfill prior months so the row is complete, not just Jul
                for m2, c2 in _month_cols(ws, header_row):
                    ws.cell(row, c2).value = round(
                        float(data["pl"].get(m2, {}).get(label, 0) or 0), 2)
                print(f"  {QBO_SHEET}/pl   r{row} '{label}' written "
                      f"(no workbook row previously existed)")
        print(f"  {QBO_SHEET}/{section:<4} col {get_column_letter(col)}: "
              f"{written} written, {missing} left blank (no JSON match)")
        for u in unmatched:
            print(f"      NO MATCH {u}")

    # 2. Consolidated P&L: translate the prior actuals column into the new one.
    ws = wb[CONSOL_SHEET]
    prev_col = find_month_col(ws, 5, prev)
    col = find_month_col(ws, 5, month)
    if prev_col is None or col is None:
        raise SystemExit(f"{CONSOL_SHEET}: could not locate '{prev}'/'{month}'")
    src_letter, dst_letter = get_column_letter(prev_col), get_column_letter(col)
    translated = 0
    for r in range(6, ws.max_row + 1):
        src = ws.cell(r, prev_col).value
        if isinstance(src, str) and src.startswith("="):
            ws.cell(r, col).value = Translator(
                src, origin=f"{src_letter}{r}").translate_formula(f"{dst_letter}{r}")
            translated += 1
    print(f"  {CONSOL_SHEET} col {dst_letter}: {translated} formulas "
          f"translated from {src_letter} (QBO refs advance automatically)")

    # 3. Studio tabs: hardcode values, translate subtotal formulas.
    for tab in STUDIO_TABS:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        code = studio_code(tab)
        month_data = data["studios"].get(code, {}).get("data", {}).get(month)
        prev_col = find_month_col(ws, 5, prev)
        col = find_month_col(ws, 5, month)
        if col is None or prev_col is None:
            print(f"  {tab:<10} column not found — skipped")
            continue
        if not month_data:
            print(f"  {tab:<10} no JSON data for {month} — left as forecast")
            continue
        src_letter = get_column_letter(prev_col)
        dst_letter = get_column_letter(col)
        vals = formulas = 0
        for r in range(6, ws.max_row + 1):
            label = ws.cell(r, 2).value
            if label is None:
                continue
            src = ws.cell(prev_col and r, prev_col).value
            if tab == "HO P&L":
                want = ho_row_value(month_data, r)
            else:
                try:
                    want = studio_row_value(month_data, label)
                except KeyError:
                    want = None
            if want is not None:
                adj = manual_adjustment(tab, r)
                ws.cell(r, col).value = round(want + adj, 2)
                if adj:
                    print(f"    {tab} r{r}: carried manual adjustment "
                          f"{adj:+,.2f} on top of {want:,.2f}")
                vals += 1
            elif isinstance(src, str) and src.startswith("="):
                ws.cell(r, col).value = Translator(
                    src, origin=f"{src_letter}{r}").translate_formula(
                        f"{dst_letter}{r}")
                formulas += 1
            ws.cell(r, col)._style = ws.cell(r, prev_col)._style
        prev_hdr = normalize(ws.cell(5, prev_col).value)
        if "(" in prev_hdr:
            marker = "(" + prev_hdr.split("(", 1)[1]
            ws.cell(5, col).value = f"{month} {marker}"
            print(f"  {tab:<10} header -> '{month} {marker}' (actuals marker)")
        print(f"  {tab:<10} col {dst_letter}: {vals} values, "
              f"{formulas} subtotal formulas")

    # 4. Sales Forecast: replace the month's forecast with Cat's actual cash
    #    sales per studio. Cat's figures are authoritative for cash sales and
    #    are already in committed_actuals (applied during the close).
    ws = wb["Sales Forecast"]
    col = find_month_col(ws, 5, month)
    dash_month = _to_dash(month)
    csf = data.get("client_sales_forecast", {})
    if col is None:
        print("  Sales Forecast: month column not found — skipped")
    else:
        total = 0.0
        # Only the monthly per-studio block (header r5 .. its TOTAL row). A
        # second "Annual Summary" table below repeats the same studio codes on
        # different columns and must not be touched.
        for r in range(6, ws.max_row + 1):
            code = normalize(ws.cell(r, 2).value)
            if code == "TOTAL":
                break
            if not code:
                continue
            val = csf.get(code, {}).get(dash_month)
            if val is None:
                continue
            ws.cell(r, col).value = round(float(val), 2)
            total += float(val)
        consol = data.get("client_sales_forecast_consolidated", {}).get(dash_month)
        print(f"  Sales Forecast col {get_column_letter(col)}: per-studio cash "
              f"sales written, sum ${total:,.0f}"
              + (f" (Cat's reported total ${consol:,.0f})" if consol else ""))

    # 5. Assumptions + header notes.
    ws = wb["Assumptions"]
    for r in range(1, ws.max_row + 1):
        if normalize(ws.cell(r, 2).value) == "Last Actuals Month":
            was = ws.cell(r, 3).value
            ws.cell(r, 3).value = month
            print(f"  Assumptions C{r}: '{was}' -> '{month}'")
            break

    ws = wb[CONSOL_SHEET]
    mon, year = month.split()
    note = (f"Actuals Jan-{mon} {year}. Forecast {_next_month(month)}+ uses "
            f"formulas from Sales Forecast and Assumptions.")
    ws.cell(3, 2).value = note
    print(f"  {CONSOL_SHEET} B3: {note}")

    if "Cover" in wb.sheetnames:
        cover = wb["Cover"]
        stamp = datetime.now().strftime("%Y-%m-%d")
        for r in range(1, min(cover.max_row, 40) + 1):
            for c in range(1, 6):
                if "refresh" in str(cover.cell(r, c).value or "").lower():
                    cover.cell(r, c + 1).value = stamp
                    print(f"  Cover: refresh date -> {stamp}")
                    break
    print("=" * 74)


def _to_dash(month: str) -> str:
    """'Jul 2026' -> '2026-07' (the key style used in client_sales_forecast)."""
    mons = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    m, y = month.split()
    return f"{y}-{mons.index(m) + 1:02d}"


def _next_month(month: str) -> str:
    mons = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    m, y = month.split()
    i = mons.index(m)
    return f"{mons[i + 1]} {y}" if i < 11 else f"Jan {int(y) + 1}"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--month", required=True, help='e.g. "Jul 2026"')
    ap.add_argument("--source", help="workbook to read")
    ap.add_argument("--out", help="workbook to write (refresh mode)")
    ap.add_argument("--validate", action="store_true",
                    help="check the row map against an existing actuals month")
    ap.add_argument("--json", help="committed_actuals to read (default: live)")
    args = ap.parse_args()

    data = json.loads(Path(args.json).read_text() if args.json
                      else DASHBOARD_JSON.read_text())
    src = Path(args.source)

    if args.validate:
        wb = load_workbook(src, data_only=True)
        raise SystemExit(1 if validate(wb, data, args.month) else 0)

    out = Path(args.out)
    shutil.copy(src, out)
    print(f"Copied source -> {out.name}\n")
    wb = load_workbook(out)  # keep formulas
    refresh(wb, data, args.month)
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
