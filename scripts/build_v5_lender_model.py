"""
v5 of the analyst's lender model.

v4 → v5 fixes Cat's 2026-06-30 P.M. flag: Consolidated P&L's
TOTAL OPERATING EXPENSES row (r21) does not equal sum(r14:r20) for the
historical actuals.

Root cause:
  (1) Consolidated P&L has no Travel & Meals row. Studio P&Ls track it
      (rows 48-50) and HO P&L tracks it (rows 28-30), but the consolidated
      r14-r20 structure skips it. ~$5K/mo leaks out.
  (2) The analyst hardcoded r21 actuals to MATCH the accountant Total
      Expenses (e.g., Jan 2026 = $945,488). The components above (r14-r20)
      sum to something less because of the missing Travel & Meals plus a
      handful of small accountant categories (parking, furniture,
      bank fees, etc.) that the model doesn't separately track.

Fix:
  (a) Repurpose r20 (currently "HO Other (Util+Fin)" — mostly zero/unused)
      to be "Travel & Meals". Formula for ALL cols (D-AM, Jan 2025 - Dec 2028)
      = SUM(studios r50 + HO r30). This captures the biggest leak.
  (b) Replace r21 TOTAL OPERATING EXPENSES with =SUM(<col>14:<col>20) for
      ALL cols (D-AM). Currently actuals are hardcoded and forecast is the
      same formula — make it consistent.
  (c) Note the residual gap: even after these fixes, the model's actuals
      r21 will be ~$5-10K/mo lower than the accountant Total Expenses for
      Jan-Apr 2026, because the model doesn't separately track 610100
      Furniture & Equipment, 611000 Shipping, 615000 Parking Lot Rental,
      613000 Bank Fees, 609000 Business licenses. These are immaterial
      (~1% of monthly OpEx) but Cat can decide whether to add new rows.
"""
from __future__ import annotations
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v4.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v5.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v5.xlsx"

# Column range: Jan 2025 (D=4) through Dec 2028 (AM=39)
JAN_2025 = 4
DEC_2028 = 39

STUDIO_TABS = ['BK P&L', 'CC P&L', 'DN P&L', 'LF P&L', 'MR P&L', 'OP P&L',
               'PH P&L', 'RH P&L', 'SB P&L', 'SM P&L', 'WP P&L', 'WW P&L']


def repurpose_r20_to_travel_meals(wb):
    """Re-label r20 from 'HO Other (Util+Fin)' to 'Travel & Meals' and set
    formula = SUM(studios r50 + HO r30) for all cols."""
    print("\n[a] Repurpose r20 → Travel & Meals across all cols (D-AM)")
    pl = wb["P&L"]
    pl.cell(row=20, column=2, value="Travel & Meals")
    pl.cell(row=20, column=2).font = Font(bold=False)
    for col in range(JAN_2025, DEC_2028 + 1):
        cl = get_column_letter(col)
        parts = [f"'{tab}'!{cl}50" for tab in STUDIO_TABS] + [f"'HO P&L'!{cl}30"]
        pl.cell(row=20, column=col).value = "=" + "+".join(parts)
    print(f"    Updated label + formula for cols D-AM")


def fix_r21_total_opex_formula(wb):
    """Replace r21 TOTAL OPERATING EXPENSES hardcoded actuals with formula.
    Same =SUM(<col>14:<col>20) the forecast already uses."""
    print("\n[b] r21 TOTAL OPERATING EXPENSES → =SUM(<col>14:<col>20) for all cols")
    pl = wb["P&L"]
    for col in range(JAN_2025, DEC_2028 + 1):
        cl = get_column_letter(col)
        pl.cell(row=21, column=col).value = f"=SUM({cl}14:{cl}20)"
    print(f"    Updated cols D-AM (now ties to sum of components by definition)")


def add_note(wb):
    """Append a note to Read Me explaining the residual accountant gap."""
    print("\n[c] Read Me note: residual OpEx categorization gap")
    rm = wb["Read Me"]
    # find first empty row after row 55 (v4 added rows 52-55)
    next_row = 57
    rm.cell(row=next_row, column=2, value="Operating Expenses — residual gap vs accountant actuals").font = Font(bold=True, size=11)
    notes = [
        ("Categories tracked",
         "The consolidated P&L tracks 7 OpEx categories: Marketing (r14), Payroll (r15), Software & Admin (r16), Professional Fees (r17), Utilities (r18), Property Costs (r19), Travel & Meals (r20)."),
        ("Categories NOT separately tracked",
         "The accountant's actuals include a handful of smaller categories the model doesn't break out: 610100 Furniture & Equipment, 611000 Shipping & postage, 613000 Bank fees (corporate), 615000 Parking Lot Rental, 609000 Business licenses, 607000 Entertainment. Combined ~$5-10K/month."),
        ("Residual gap",
         "Model's r21 actuals ≈ accountant 'Total Expenses' minus the ~$5-10K of untracked categories. Material for accountant reconciliation, immaterial (~1% of monthly OpEx) for the lender forecast model. To close the gap, add new rows on P&L + studio/HO P&L tabs to capture these categories explicitly."),
    ]
    for i, (label, body) in enumerate(notes, start=1):
        rm.cell(row=next_row + i, column=2, value=label).font = Font(bold=True)
        rm.cell(row=next_row + i, column=3, value=body).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"    Added rows {next_row}-{next_row + len(notes)}")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v4 file missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v4 → {OUT.name}")

    wb = load_workbook(OUT)
    repurpose_r20_to_travel_meals(wb)
    fix_r21_total_opex_formula(wb)
    add_note(wb)

    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
