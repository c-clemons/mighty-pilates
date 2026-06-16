"""
Generate Saasant-format MTT geographic reclass JEs for Feb, Mar, May 2026.

Cat's policy (2026-06-16):
  Revenue follows session location, not sale location.
    Marin       <- Berkeley, Lafayette, Russian Hill, Presidio Heights, Danville
    Westwood    <- Culver City, Santa Monica, Ocean Park
    Santa Barbara stays Santa Barbara
    Marin / Westwood / Santa Barbara own amounts stay put.

Net amount per JE = $0 (pure reclass). Total MTT per month unchanged in QBO.

Inputs: QBO P&L by Location exports for Feb / Mar / May 2026.
Output: outputs/MTT_Reclass_Feb-Mar-May_2026_<stamp>.xlsx
"""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

INPUT_FILES = {
    ("Feb 2026", 2, "2026-02-28", "26.02"): (
        "/Users/chandlerclemons/Downloads/Norbrook Lifestyle LLC_Profit and Loss by Location Feb.xlsx"
    ),
    ("Mar 2026", 3, "2026-03-31", "26.03"): (
        "/Users/chandlerclemons/Downloads/Norbrook Lifestyle LLC_Profit and Loss by Location March.xlsx"
    ),
    ("May 2026", 5, "2026-05-31", "26.05"): (
        "/Users/chandlerclemons/Downloads/Norbrook Lifestyle LLC_Profit and Loss by Location May.xlsx"
    ),
}

# Mapping: where each studio's MTT revenue should land (session location)
DESTINATION = {
    "Berkeley":         "Marin",
    "Lafayette":        "Marin",
    "Russian Hill":     "Marin",
    "Presidio Heights": "Marin",
    "Danville":         "Marin",
    "Culver City":      "Westwood",
    "Santa Monica":     "Westwood",
    "Ocean Park":       "Westwood",
    "Marin":            "Marin",          # own — no move
    "Westwood":         "Westwood",       # own — no move
    "Santa Barbara":    "Santa Barbara",  # own — no move
}

ACCOUNT_NAME = "Mighty Teacher Training"  # QBO account name on GL 401004


def read_mtt_by_studio(path: str) -> dict:
    """
    Read MTT row from a QBO P&L by Location export.
    Returns {studio_name: amount} for studios with non-zero MTT.
    """
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Header row (5) has studio names; MTT row matches '401004' or 'teacher training'
    rows = list(ws.iter_rows(values_only=True))
    header = rows[4]  # row 5 → index 4

    mtt_row = None
    for row in rows:
        if row and row[0] and "401004" in str(row[0]):
            mtt_row = row
            break
    if mtt_row is None:
        raise RuntimeError(f"No MTT row in {path}")

    # Pair each header cell with the MTT cell, skip None, 'Total', 'Total for Marin',
    # 'Head Office', 'West Portal', 'Not specified' (placeholder studios)
    out = {}
    SKIP = {"", "Head Office", "West Portal", "Not specified", "Total", None}
    seen_marin = False  # 'Marin' appears twice in some exports
    for h, v in zip(header, mtt_row):
        if h in SKIP:                          continue
        if isinstance(h, str) and h.startswith("Total"):  continue
        if v is None or float(v) == 0:        continue
        if h == "Marin":
            if seen_marin: continue
            seen_marin = True
        out[h] = float(v)
    return out


def build_reclass_lines(mtt_by_studio: dict) -> list:
    """
    Given {studio: amount}, return list of (location, signed_amount) tuples.
    Source studios: positive amount (debit — removes revenue from source).
    Destination studios: negative amount (credit — adds revenue at destination).
    Net = 0.
    """
    debits  = []   # (location, +amount) — source debits
    credits = {}   # destination -> total credit amount

    for studio, amt in mtt_by_studio.items():
        dest = DESTINATION.get(studio)
        if dest is None:
            raise RuntimeError(f"No destination mapping for studio {studio!r}")
        if dest == studio:
            continue  # own revenue, no move
        debits.append((studio, amt))
        credits[dest] = credits.get(dest, 0.0) + amt

    lines = []
    for studio, amt in sorted(debits):
        lines.append((studio, round(amt, 2)))                    # positive = debit
    for dest, amt in sorted(credits.items()):
        lines.append((dest, -round(amt, 2)))                     # negative = credit
    return lines


def write_workbook(per_month_lines: dict, output_path: str):
    """
    Write Saasant-format Excel with one JE per month.
    Columns match the existing Saasant convention from pipeline/saasant_export.py.
    """
    headers = [
        "Journal No", "Journal Date", "Memo", " Account ", " Amount", " Description",
        "Name", "Location", "Class ", "Currency Code", "Exchange Rate", "Is Adjustment",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Journal Entries"
    ws.append(headers)

    row_idx = 2
    for (label, date_str, journal_no, memo, description), lines in per_month_lines.items():
        first_excel_row = row_idx
        for i, (location, amount) in enumerate(lines):
            if i == 0:
                # First row: full Journal No, Date, Description
                ws.append([
                    journal_no, date_str, memo, ACCOUNT_NAME, amount,
                    description, None, location, None, None, None, None,
                ])
            else:
                # Subsequent rows: reference back to row N
                ws.append([
                    f"=A{first_excel_row}", None, None, ACCOUNT_NAME, amount,
                    f"=F{first_excel_row}", None, location, None, None, None, None,
                ])
            row_idx += 1

    wb.save(output_path)


def main():
    per_month_lines = {}
    print("=== MTT geographic reclass — Feb/Mar/May 2026 ===\n")

    for (label, month_num, date_str, yy_mm), path in INPUT_FILES.items():
        mtt = read_mtt_by_studio(path)
        lines = build_reclass_lines(mtt)
        gross = sum(abs(x[1]) for x in lines) / 2  # sum of moves
        net   = sum(x[1] for x in lines)

        print(f"--- {label} ({date_str}) ---")
        print(f"  QBO MTT total: ${sum(mtt.values()):,.2f}")
        print(f"  Gross moves:   ${gross:,.2f}")
        print(f"  Net (should=0): ${net:,.2f}")
        print(f"  JE lines:")
        for studio, amt in lines:
            sign = "DR" if amt > 0 else "CR"
            print(f"    {sign}  {studio:<22}  {amt:>+12,.2f}")
        print()

        per_month_lines[(
            label,
            date_str,
            f"MTT Reclass {yy_mm}",
            None,  # memo — leave blank
            f"MTT geographic reclass — session location (Cat directive 2026-06-16) — {label}",
        )] = lines

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = Path("/Users/chandlerclemons/mighty-pilates/outputs") / (
        f"MTT_Reclass_Feb-Mar-May_2026_{stamp}.xlsx"
    )
    out_path.parent.mkdir(exist_ok=True)
    write_workbook(per_month_lines, str(out_path))
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
