"""
v8: fix circular reference in WW P&L r68 → r69 → r68.

v7 introduced r69 = SUM(r65:r68) across all studio P&L tabs (to fix per-studio
Net Income). But Westwood's r68 Property Taxes forecast still had the analyst's
legacy formula =<col>69*0.658496 (allocating 65.85% of Total Other Expenses to
Property Taxes). Combined with the new r69 formula, this creates a cycle:

  WW r68 = r69 * 0.658496
  WW r69 = SUM(r65:r68)

The 11 other studios had r68 hardcoded to 0 already. Only WW had the formula
because the analyst assigned all consolidated Property Taxes to the Westwood
ledger.

Fix: zero out WW r68 forecast (cols U-AM). Actuals cols D-T retain Westwood's
$3,673/mo hardcoded value. Consolidated Property Taxes (P&L r26) is hardcoded
at $4,178.19/mo per v4 — doesn't sum from studios, so this change doesn't
disturb the consolidated total.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v7.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v8.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v8.xlsx"


def zero_ww_property_tax_forecast(wb):
    print("\n[a] Zero WW P&L r68 Property Taxes forecast (cols U-AM)")
    ww = wb["WW P&L"]
    for col in range(21, 40):  # U-AM = Jun 2026 - Dec 2028
        ww.cell(row=68, column=col).value = 0
    print(f"    Overwrote legacy circular formula =<col>69*0.658496 with 0")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v7 missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v7 → {OUT.name}")

    wb = load_workbook(OUT)
    zero_ww_property_tax_forecast(wb)
    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
