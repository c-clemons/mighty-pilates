"""
v6: rebuild the consolidated P&L 2025 actuals (cols D-O) from QBO source +
make every Total row a SUM formula so they tie by definition.

Cat 2026-06-30 flag: "Total rows don't sum properly... values need to match
the actuals from QBO." Root cause:

  (1) Model's r11 Total Revenue 2025 actuals were hardcoded to QBO 'Total
      Income' which includes 402000 Revenue from Old Mighty (~$180K Jan
      tailing to $1K Dec) — but no row on the consolidated P&L captures
      this category. Sum(r6:r10) is short by the Old Mighty amount.

  (2) Other QBO categories absent from the model: 607000 Entertainment,
      608000 Insurance (HO portion), 609000 Business Licenses, 610100
      Furniture & Equipment, 611000 Shipping & Postage, 613000 Bank Fees,
      615000 Parking Lot Rental, 630000 Studio Start Up Costs.

Fix:
  (a) For 2025 cols D-O, REBUILD every consolidated P&L row from QBO 2025
      file directly. r6 Session Revenue absorbs Old Mighty (combined).
      r16 Software absorbs Insurance + Office Supplies + Furniture +
      Shipping + Bank Fees ("Software & Admin overhead"). r20 absorbs
      Entertainment + Business Licenses + Parking + Studio Start Up
      ("Other Operating"). All other rows: direct QBO mapping.
  (b) Replace every Total row formula with =SUM range for all cols:
      r11 Total Revenue   = SUM(r6:r10)
      r13 Gross Profit    = r11 - r12
      r22 NOI             = r13 - r21
      r27 Total Other Exp = SUM(r23:r26)
      r28 Net Income      = r22 - r27
      (r21 Total OpEx = SUM(r14:r20) already from v5)
  (c) Relabel r20 from "Travel & Meals" to "Other Operating Expenses"
      since it now catches more than just T&M for 2025. For 2026+ forecast,
      r20 still reads SUM(studios r50 + HO r30) — gives T&M only — that's
      a conservative forecast (Other items aren't separately forecast).
  (d) Add Read Me note explaining the 2025/2026+ category mapping.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v5.xlsx")
QBO_2025 = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Pilates_Financials_2025_062326.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v6.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v6.xlsx"

# Column extent
JAN_2025 = 4    # D
DEC_2025 = 15   # O
DEC_2028 = 39   # AM

# QBO row mapping (verified from inspection of Mighty Pilates_Financials_2025_062326.xlsx)
# Row index in QBO file
Q_TOTAL_SESSIONS = 14         # Total 401000 Sessions
Q_OLD_MIGHTY = 15             # 402000 Revenue from Old Mighty
Q_TOTAL_BREAKAGE = 21         # Total 403000 Breakage Revenue
Q_RETAIL = 22                 # 404000 Retail Sales
Q_REFUNDS = 23                # 406000 Refunds (negative in QBO)
Q_DISCOUNTS = 24              # 407000 Discounts (negative in QBO)
Q_TOTAL_COGS = 32             # Total Cost of Goods Sold
Q_TOTAL_MARKETING = 42        # Total 601000 Sales & Marketing
Q_TOTAL_PAYROLL = 51          # Total 602000 Payroll
Q_603_SOFTWARE = 52           # 603000 Software & Web Services
Q_TOTAL_PROFESSIONAL = 58     # Total 604000 Professional Fees
Q_605_TRAVEL = 59
Q_606_MEALS = 60
Q_607_ENTERTAIN = 61
Q_608_INSURANCE = 62
Q_609_LICENSES = 63
Q_610_OFFICE = 64
Q_610100_FURNITURE = 65
Q_611_SHIPPING = 66
Q_613_BANK = 67
Q_615_PARKING = 68
Q_TOTAL_UTILITIES = 75        # Total 616000 Utilities
Q_630_START_UP = 76           # Studio Start Up Costs
Q_TOTAL_PROPERTY = 83         # Total 700000 Property Costs
Q_810_DEPRECIATION = 87
Q_900_OTHER = 88              # 900000 Other Expense/(Income)
Q_901_INTEREST = 89
Q_902_TAXES = 90
Q_903_PROPERTY_TAX = 91

# Consolidated P&L row mapping
M_SESSION_REVENUE = 6
M_BREAKAGE = 7
M_RETAIL = 8
M_REFUNDS = 9
M_DISCOUNTS = 10
M_TOTAL_REVENUE = 11
M_COGS = 12
M_GROSS_PROFIT = 13
M_MARKETING = 14
M_PAYROLL = 15
M_SOFTWARE = 16
M_PROFESSIONAL = 17
M_UTILITIES = 18
M_PROPERTY = 19
M_OTHER_OPERATING = 20  # was Travel & Meals after v5
M_TOTAL_OPEX = 21
M_NOI = 22
M_DEPRECIATION = 23
M_INTEREST = 24
M_TAXES_PAID = 25
M_PROPERTY_TAXES = 26
M_TOTAL_OTHER = 27
M_NET_INCOME = 28


def q(qbo_ws, row, qbo_col):
    v = qbo_ws.cell(row=row, column=qbo_col).value
    return float(v) if isinstance(v, (int, float)) else 0.0


def rebuild_2025_consolidated_pl(wb, qbo_ws):
    """For each col D-O (Jan-Dec 2025), write QBO-sourced values to the
    consolidated P&L row mapping. Sums tie by formula."""
    print("\n[a] Rebuilding 2025 consolidated P&L actuals from QBO (cols D-O)")
    pl = wb["P&L"]

    # QBO cols B-M = Jan-Dec 2025 (cols 2-13)
    # Model cols D-O = Jan-Dec 2025 (cols 4-15)
    # Offset: model col = qbo col + 2

    for model_col in range(JAN_2025, DEC_2025 + 1):
        qbo_col = model_col - 2  # B=2 for Jan 2025

        # Revenue rows
        sessions = q(qbo_ws, Q_TOTAL_SESSIONS, qbo_col)
        old_mighty = q(qbo_ws, Q_OLD_MIGHTY, qbo_col)
        pl.cell(row=M_SESSION_REVENUE, column=model_col).value = sessions + old_mighty
        pl.cell(row=M_BREAKAGE, column=model_col).value = q(qbo_ws, Q_TOTAL_BREAKAGE, qbo_col)
        pl.cell(row=M_RETAIL, column=model_col).value = q(qbo_ws, Q_RETAIL, qbo_col)
        pl.cell(row=M_REFUNDS, column=model_col).value = q(qbo_ws, Q_REFUNDS, qbo_col)
        pl.cell(row=M_DISCOUNTS, column=model_col).value = q(qbo_ws, Q_DISCOUNTS, qbo_col)

        # COGS
        pl.cell(row=M_COGS, column=model_col).value = q(qbo_ws, Q_TOTAL_COGS, qbo_col)

        # OpEx components
        pl.cell(row=M_MARKETING, column=model_col).value = q(qbo_ws, Q_TOTAL_MARKETING, qbo_col)
        pl.cell(row=M_PAYROLL, column=model_col).value = q(qbo_ws, Q_TOTAL_PAYROLL, qbo_col)
        # Software & Admin = 603 + 608 + 610 + 610100 + 611 + 613
        software_admin = (
            q(qbo_ws, Q_603_SOFTWARE, qbo_col) +
            q(qbo_ws, Q_608_INSURANCE, qbo_col) +
            q(qbo_ws, Q_610_OFFICE, qbo_col) +
            q(qbo_ws, Q_610100_FURNITURE, qbo_col) +
            q(qbo_ws, Q_611_SHIPPING, qbo_col) +
            q(qbo_ws, Q_613_BANK, qbo_col)
        )
        pl.cell(row=M_SOFTWARE, column=model_col).value = software_admin
        pl.cell(row=M_PROFESSIONAL, column=model_col).value = q(qbo_ws, Q_TOTAL_PROFESSIONAL, qbo_col)
        pl.cell(row=M_UTILITIES, column=model_col).value = q(qbo_ws, Q_TOTAL_UTILITIES, qbo_col)
        pl.cell(row=M_PROPERTY, column=model_col).value = q(qbo_ws, Q_TOTAL_PROPERTY, qbo_col)

        # Other Operating = 605 + 606 + 607 + 609 + 615 + 630
        other_op = (
            q(qbo_ws, Q_605_TRAVEL, qbo_col) +
            q(qbo_ws, Q_606_MEALS, qbo_col) +
            q(qbo_ws, Q_607_ENTERTAIN, qbo_col) +
            q(qbo_ws, Q_609_LICENSES, qbo_col) +
            q(qbo_ws, Q_615_PARKING, qbo_col) +
            q(qbo_ws, Q_630_START_UP, qbo_col)
        )
        pl.cell(row=M_OTHER_OPERATING, column=model_col).value = other_op

        # Other Expenses — combine 900000 Other Income/Expense into r24 Interest Expense
        pl.cell(row=M_DEPRECIATION, column=model_col).value = q(qbo_ws, Q_810_DEPRECIATION, qbo_col)
        pl.cell(row=M_INTEREST, column=model_col).value = (
            q(qbo_ws, Q_900_OTHER, qbo_col) + q(qbo_ws, Q_901_INTEREST, qbo_col)
        )
        pl.cell(row=M_TAXES_PAID, column=model_col).value = q(qbo_ws, Q_902_TAXES, qbo_col)
        pl.cell(row=M_PROPERTY_TAXES, column=model_col).value = q(qbo_ws, Q_903_PROPERTY_TAX, qbo_col)

    print(f"    Populated 2025 cols D-O for revenue, COGS, 7 OpEx components, 4 Other Expense rows")


def set_total_formulas(wb):
    """Replace hardcoded Total rows with =SUM formulas for ALL cols.
    Ensures Totals tie to components by definition."""
    print("\n[b] Total rows → =SUM formulas (all cols D-AM)")
    pl = wb["P&L"]
    for col in range(JAN_2025, DEC_2028 + 1):
        cl = get_column_letter(col)
        pl.cell(row=M_TOTAL_REVENUE, column=col).value = f"=SUM({cl}{M_SESSION_REVENUE}:{cl}{M_DISCOUNTS})"
        pl.cell(row=M_GROSS_PROFIT, column=col).value = f"={cl}{M_TOTAL_REVENUE}-{cl}{M_COGS}"
        pl.cell(row=M_NOI, column=col).value = f"={cl}{M_GROSS_PROFIT}-{cl}{M_TOTAL_OPEX}"
        pl.cell(row=M_TOTAL_OTHER, column=col).value = f"=SUM({cl}{M_DEPRECIATION}:{cl}{M_PROPERTY_TAXES})"
        pl.cell(row=M_NET_INCOME, column=col).value = f"={cl}{M_NOI}-{cl}{M_TOTAL_OTHER}"
    print(f"    r{M_TOTAL_REVENUE} Total Revenue = SUM(r{M_SESSION_REVENUE}:r{M_DISCOUNTS})")
    print(f"    r{M_GROSS_PROFIT} Gross Profit  = r{M_TOTAL_REVENUE} - r{M_COGS}")
    print(f"    r{M_NOI} NOI            = r{M_GROSS_PROFIT} - r{M_TOTAL_OPEX}")
    print(f"    r{M_TOTAL_OTHER} Total Other Exp= SUM(r{M_DEPRECIATION}:r{M_PROPERTY_TAXES})")
    print(f"    r{M_NET_INCOME} Net Income     = r{M_NOI} - r{M_TOTAL_OTHER}")


def relabel_other_operating(wb):
    """Update r20 label since it now catches more than just Travel & Meals."""
    print("\n[c] Relabel r20 → 'Other Operating Expenses'")
    pl = wb["P&L"]
    pl.cell(row=M_OTHER_OPERATING, column=2).value = "Other Operating Expenses"


def add_note(wb):
    """Read Me note explaining the 2025/2026+ category mapping."""
    print("\n[d] Read Me note: 2025 vs forecast category mapping")
    rm = wb["Read Me"]
    start = 62
    rm.cell(row=start, column=2, value="Consolidated P&L category mapping").font = Font(bold=True, size=11)
    notes = [
        ("Why 2025 is hardcoded from QBO",
         "2025 consolidated P&L values are pulled directly from the QBO 2025 financial package (Mighty Pilates_Financials_2025_062326.xlsx) to guarantee tie-out to accountant. They are not summed from the per-studio P&L tabs."),
        ("Where Old Mighty Revenue lives",
         "QBO has a 402000 Revenue from Old Mighty line ($440K in 2025, tailing from $180K Jan to $1K Dec). It is absorbed into r6 Session Revenue for 2025 (no separate row). For 2026+ it is $0."),
        ("Row 16 Software composition",
         "For 2025 actuals, r16 covers QBO 603 Software + 608 Insurance + 610 Office Supplies + 610100 Furniture & Equipment + 611 Shipping & Postage + 613 Bank Fees (corporate). Best read as 'Software & Admin Overhead'."),
        ("Row 20 Other Operating composition",
         "For 2025 actuals, r20 covers QBO 605 Travel + 606 Meals + 607 Entertainment + 609 Business Licenses + 615 Parking Lot Rental + 630 Studio Start Up Costs. For 2026+ forecast, r20 formula sums Travel & Meals only (studios r50 + HO r30) — conservative forecast that excludes one-time start-up + parking."),
        ("All Total rows are now formulas",
         "r11 Total Revenue, r13 Gross Profit, r22 NOI, r27 Total Other Expenses, r28 Net Income all use =SUM or arithmetic formulas now. Sums tie to components by definition."),
    ]
    for i, (label, body) in enumerate(notes, start=1):
        rm.cell(row=start + i, column=2, value=label).font = Font(bold=True)
        rm.cell(row=start + i, column=3, value=body).alignment = Alignment(wrap_text=True, vertical="top")
    print(f"    Added rows {start}-{start + len(notes)}")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v5 missing: {SRC}")
    if not QBO_2025.exists():
        raise FileNotFoundError(f"QBO 2025 file missing: {QBO_2025}")

    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v5 → {OUT.name}")

    wb = load_workbook(OUT)
    qbo = load_workbook(QBO_2025, data_only=True)
    qbo_ws = qbo["PL"]

    rebuild_2025_consolidated_pl(wb, qbo_ws)
    set_total_formulas(wb)
    relabel_other_operating(wb)
    add_note(wb)

    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
