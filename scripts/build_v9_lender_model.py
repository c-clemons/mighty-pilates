"""
v9: eliminate ALL remaining circular refs from studio P&L r66 Interest Expense.

v8 fixed WW r68 Property Taxes but missed the same legacy pattern in r66
Interest Expense for 4 studios:

  LF r66 U-AM: =<col>69*0.259451
  MR r66 U-AM: =<col>69*0.395554
  PH r66 U-AM: =<col>69*0.725744
  SM r66 U-AM: =<col>69*0.602586

Combined with v7's r69=SUM(r65:r68) this creates 76 individual r66↔r69 cycles
across those 4 studios × 19 forecast cols (Jun 2026 - Dec 2028).

Fix: zero out r66 forecast cols U-AM (21-39) for LF/MR/PH/SM. Actuals cols
D-T retain the analyst's hardcoded values ($1,397 / $1,366 / $2,405 / $3,890
for Jan 2026, etc.). Consolidated Interest Expense is already handled at
HO P&L r39 which links to Cash Flow Forecast!49 (debt schedule) plus PF Fees
amortization from Sept 2026 forward — that's the single source of truth for
forecast interest per v3.
"""
from __future__ import annotations
import shutil
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
SRC = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v8.xlsx")
OUT = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Updated DRAFT_External_Mighty Pilates_Financial Workbook_6.25.2026_v9.xlsx")
SNAP = REPO / "snapshots" / "excel" / "Updated_DRAFT_External_Mighty_Pilates_Financial_Workbook_6.25.2026_v9.xlsx"

STUDIOS_WITH_LEGACY_R66 = ["LF P&L", "MR P&L", "PH P&L", "SM P&L"]


def zero_studio_interest_forecast(wb):
    print("\n[a] Zero r66 Interest Expense forecast (cols U-AM) for LF/MR/PH/SM")
    for tab in STUDIOS_WITH_LEGACY_R66:
        ws = wb[tab]
        for col in range(21, 40):  # U-AM
            ws.cell(row=66, column=col).value = 0
        print(f"    {tab}: zeroed 19 forecast cols")


def main():
    if not SRC.exists():
        raise FileNotFoundError(f"v8 missing: {SRC}")
    SNAP.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(SRC, OUT)
    print(f"Copied v8 → {OUT.name}")
    wb = load_workbook(OUT)
    zero_studio_interest_forecast(wb)
    wb.save(OUT)
    shutil.copy(OUT, SNAP)
    print(f"\nSaved:    {OUT}")
    print(f"Snapshot: {SNAP.relative_to(REPO)}")


if __name__ == "__main__":
    main()
