"""
Apply Cat's authoritative June 2026 cash sales per studio.

Source: Mighty June Sales Data.csv (2026-07-07)
Total: $759,206 (per-studio sum: $759,205, $1 rounding)

Updates:
  1. dashboard/data/baseline.json           (git-tracked persistent state)
  2. dashboard/data/user_overrides.json     (Cash Flow page reads this)
  3. dashboard/data/committed_actuals.json  (client_sales_forecast tree)
  4. Excel Sales Forecast tab, col I = Jun 2026 (A)
  5. Excel Cash Flow Forecast row 7 → formula reference to Sales Forecast
"""
from __future__ import annotations
import json
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

CAT_JUN_2026 = {
    "BK": 100821,
    "CC":  32522,
    "DN":  20327,
    "LF":  81024,
    "MR":  98742,
    "OP":  50286,
    "PH": 113383,
    "RH":  71619,
    "SB":  17783,
    "SM": 130716,
    "WP":    225,
    "WW":  41757,
}
CAT_TOTAL = sum(CAT_JUN_2026.values())    # 759,205
TARGET_TOTAL = 759206                     # CSV grand total row

MONTH_DASH = "2026-06"
EXCEL_COL_INDEX = 9  # I = Jun 2026 (A)

REPO_ROOT = Path(__file__).resolve().parents[1]
BASELINE_JSON = REPO_ROOT / "dashboard" / "data" / "baseline.json"
USER_OVERRIDES_JSON = REPO_ROOT / "dashboard" / "data" / "user_overrides.json"
DASHBOARD_JSON = REPO_ROOT / "dashboard" / "data" / "committed_actuals.json"

EXCEL_LIVE_PATH = Path(
    "/Users/chandlerclemons/Desktop/Empirica Financial Modeling/Mighty Pilates/"
    "Mighty Pilates Financial Model.xlsx"
)
EXCEL_SECONDARY_PATH = Path("/Users/chandlerclemons/Desktop/Mighty Pilates/Mighty Pilates Financial Model.xlsx")
EXCEL_REPO_SNAPSHOT = REPO_ROOT / "snapshots" / "excel" / "Mighty_Pilates_Financial_Model_Jun2026.xlsx"


def update_baseline():
    print(f"\n[A0] Updating {BASELINE_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(BASELINE_JSON.read_text())
    sf = data.setdefault("sales_forecast", {})
    for code, amt in CAT_JUN_2026.items():
        bucket = sf.setdefault(code, {})
        old = bucket.get(MONTH_DASH, 0)
        bucket[MONTH_DASH] = float(amt)
        print(f"   baseline[{code}][{MONTH_DASH}]: ${old:>12,.2f} → ${amt:>12,.2f}")
    BASELINE_JSON.write_text(json.dumps(data, indent=2))


def update_user_overrides():
    print(f"\n[A] Updating {USER_OVERRIDES_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(USER_OVERRIDES_JSON.read_text())
    sf = data.setdefault("sales_forecast", {})
    for code, amt in CAT_JUN_2026.items():
        bucket = sf.setdefault(code, {})
        old = bucket.get(MONTH_DASH, 0)
        bucket[MONTH_DASH] = float(amt)
        print(f"   user_overrides[{code}][{MONTH_DASH}]: ${old:>12,.2f} → ${amt:>12,.2f}")
    data["_last_updated"] = datetime.now().isoformat()
    USER_OVERRIDES_JSON.write_text(json.dumps(data, indent=2))


def update_dashboard_json():
    print(f"\n[1/4] Updating {DASHBOARD_JSON.relative_to(REPO_ROOT)}")
    data = json.loads(DASHBOARD_JSON.read_text())
    csf = data.setdefault("client_sales_forecast", {})
    csfc = data.setdefault("client_sales_forecast_consolidated", {})
    ms = data.setdefault("monthly_sales", {})
    for code, amt in CAT_JUN_2026.items():
        bucket = csf.setdefault(code, {})
        bucket[MONTH_DASH] = float(amt)
    csfc[MONTH_DASH] = float(TARGET_TOTAL)
    ms[MONTH_DASH] = float(TARGET_TOTAL)
    print(f"   client_sales_forecast_consolidated[{MONTH_DASH}] = ${TARGET_TOTAL:,}")
    print(f"   monthly_sales[{MONTH_DASH}]                     = ${TARGET_TOTAL:,}")
    data["_last_updated"] = datetime.now().isoformat()
    DASHBOARD_JSON.write_text(json.dumps(data, indent=2))


def update_excel():
    print(f"\n[2/4] Updating Excel: {EXCEL_LIVE_PATH.name}")
    if not EXCEL_LIVE_PATH.exists():
        raise FileNotFoundError(f"Live Excel not found: {EXCEL_LIVE_PATH}")
    wb = load_workbook(EXCEL_LIVE_PATH)

    sf = wb["Sales Forecast"]
    code_to_row = {}
    for r in range(6, 18):
        code = sf.cell(row=r, column=2).value
        if code:
            code_to_row[str(code).strip()] = r

    for code, amt in CAT_JUN_2026.items():
        if code not in code_to_row:
            print(f"   WARNING: studio code {code!r} not found in Sales Forecast tab")
            continue
        r = code_to_row[code]
        cell = sf.cell(row=r, column=EXCEL_COL_INDEX)
        old = cell.value
        cell.value = float(amt)
        old_str = f"${float(old):>12,.2f}" if isinstance(old, (int, float)) else f"{old}"
        print(f"   SalesForecast!I{r}  ({code}): {old_str:>16} → ${amt:>12,.2f}")

    sum_cell = sf.cell(row=18, column=EXCEL_COL_INDEX)
    print(f"   SalesForecast!I18 (TOTAL formula): {sum_cell.value}")

    # Cash Flow Forecast row 7, col I = Jun 2026 — formula reference
    cf = wb["Cash Flow Forecast"]
    cf_cell = cf.cell(row=7, column=EXCEL_COL_INDEX)
    old_cf = cf_cell.value
    cf_cell.value = "='Sales Forecast'!I18"
    print(f"\n   CashFlowForecast!I7 Total Cash Sales: {old_cf!r} → ='Sales Forecast'!I18")

    # Update header note to reflect June is now actuals
    explain_cell = sf.cell(row=3, column=2)
    val = str(explain_cell.value or "")
    if "May 2026" in val and "Jun-" not in val and "Jun 2026" not in val:
        explain_cell.value = (
            "Jan-Jun 2026 actuals (locked, Cat-authoritative). "
            "Jul 2026 onwards forecast (editable)."
        )
        print(f"   SalesForecast!B3 header → 'Jan-Jun 2026 actuals (locked)...'")

    return wb


def save_excel_everywhere(wb):
    print(f"\n[3/4] Saving Excel to 3 locations")
    for p in [EXCEL_LIVE_PATH, EXCEL_SECONDARY_PATH, EXCEL_REPO_SNAPSHOT]:
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        print(f"   ✓ {p}")


def main():
    print("=" * 78)
    print(f"Apply Cat's June 2026 cash sales (total ${TARGET_TOTAL:,})")
    print("=" * 78)
    update_baseline()
    update_user_overrides()
    update_dashboard_json()
    wb = update_excel()
    save_excel_everywhere(wb)
    print(f"\n[4/4] Done.")
    print(f"      Per-studio sum: ${CAT_TOTAL:,}")
    print(f"      Reported total: ${TARGET_TOTAL:,}")
    print(f"      Rounding delta: ${TARGET_TOTAL - CAT_TOTAL}")


if __name__ == "__main__":
    main()
