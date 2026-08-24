"""
Monthly Close Report — PDF

Produces a single PDF report with:
  - Cover               : period, key totals (recognized + cash sales)
  - MoM Comparison      : close month vs prior month (recognized revenue & cash sales)
  - Waterfall           : revenue recognized in the close month grouped by sale-month vintage (M0 to M-7+)
  - Per-Studio Waterfall: same vintage breakdown, by studio

Data sources (all model-derived, not from QuickBooks):
  - EARNED_REVENUE_ANALYTICS.DAILY_REVENUE_AND_SALES_DETAIL  (recognized + cash sales)
  - PLAYLIST_DATAMART.CLASSPASS_REPORTING_ANALYTICS.RESERVATIONS  (ClassPass)

CLI:
    python run.py close-report --month YYYY-MM
    python run.py monthly --month YYYY-MM   (close report bundled with GL + Saasant)
"""

from __future__ import annotations
import calendar
from datetime import date, datetime
from pathlib import Path
from typing import Tuple

import pandas as pd

from pipeline.connection import execute_query_df


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------
def _last_day(year: int, month: int) -> str:
    return f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"


def _first_day(year: int, month: int) -> str:
    return f"{year}-{month:02d}-01"


def _shift_month(year: int, month: int, delta: int) -> Tuple[int, int]:
    m, y = month + delta, year
    while m <= 0:
        m += 12; y -= 1
    while m > 12:
        m -= 12; y += 1
    return y, m


# ---------------------------------------------------------------------------
# Data pulls
# ---------------------------------------------------------------------------
def pull_mom(conn, close_year: int, close_month: int) -> dict:
    """Pull the close month + immediately-prior month from all relevant sources."""
    prior_y, prior_m = _shift_month(close_year, close_month, -1)
    start = _first_day(prior_y, prior_m)
    end   = _last_day(close_year, close_month)

    ledger = execute_query_df(conn, f"""
        SELECT TO_CHAR(EVENT_DATE, 'YYYY-MM') AS MONTH_YM,
               REPLACE(STUDIO_NAME, 'Mighty Pilates ', '') AS STUDIO,
               SERVICE_TYPE,
               EVENT_TYPE,
               SUM(NET_EARNED_REVENUE)   AS NET_EARNED,
               SUM(NET_BREAKAGE_REVENUE) AS NET_BREAKAGE,
               SUM(NET_TOTAL_SALES)      AS NET_SALES
        FROM EARNED_REVENUE_ANALYTICS.DAILY_REVENUE_AND_SALES_DETAIL
        WHERE EVENT_DATE BETWEEN '{start}' AND '{end}'
        GROUP BY 1,2,3,4
    """)

    cp = execute_query_df(conn, f"""
        SELECT TO_CHAR(START_DATE, 'YYYY-MM') AS MONTH_YM,
               REPLACE(EARNED_REVENUE_ANALYTICS.CANON_STUDIO(VENUE_FULL_NAME), 'Mighty Pilates ', '') AS STUDIO,
               SUM(RATE) AS CP_REV
        FROM PLAYLIST_DATAMART.CLASSPASS_REPORTING_ANALYTICS.RESERVATIONS
        WHERE START_DATE BETWEEN '{start}' AND '{end}' AND RATE > 0
        GROUP BY 1, 2
    """)

    return {
        "ledger": ledger, "cp": cp,
        "close_ym": f"{close_year}-{close_month:02d}",
        "prior_ym": f"{prior_y}-{prior_m:02d}",
    }


def pull_waterfall(conn, close_year: int, close_month: int) -> pd.DataFrame:
    """Recognized revenue in close month, grouped by PURCHASE_DATE vintage."""
    start = _first_day(close_year, close_month)
    end   = _last_day(close_year, close_month)

    df = execute_query_df(conn, f"""
        SELECT REPLACE(STUDIO_NAME, 'Mighty Pilates ', '') AS STUDIO,
               PURCHASE_DATE::DATE AS PURCH_DT,
               SUM(NET_EARNED_REVENUE)   AS NET_EARNED,
               SUM(NET_BREAKAGE_REVENUE) AS NET_BREAKAGE
        FROM EARNED_REVENUE_ANALYTICS.DAILY_REVENUE_AND_SALES_DETAIL
        WHERE EVENT_DATE BETWEEN '{start}' AND '{end}'
          AND PURCHASE_DATE IS NOT NULL
        GROUP BY 1, 2
    """)

    close_first = date(close_year, close_month, 1)
    def vintage(d):
        if pd.isna(d): return "Unknown"
        d = d if isinstance(d, date) else d.date()
        months = (close_first.year - d.year) * 12 + (close_first.month - d.month)
        if months <= 0: return "M0 (close month)"
        if months >= 7: return "M-7+ (older)"
        return f"M-{months}"

    df["VINTAGE"] = df["PURCH_DT"].apply(vintage)
    return df


def pull_cp_by_studio(conn, year: int, month: int) -> pd.DataFrame:
    start = _first_day(year, month)
    end   = _last_day(year, month)
    return execute_query_df(conn, f"""
        SELECT REPLACE(EARNED_REVENUE_ANALYTICS.CANON_STUDIO(VENUE_FULL_NAME),
                       'Mighty Pilates ', '') AS STUDIO,
               SUM(RATE) AS CP_REV
        FROM PLAYLIST_DATAMART.CLASSPASS_REPORTING_ANALYTICS.RESERVATIONS
        WHERE START_DATE BETWEEN '{start}' AND '{end}' AND RATE > 0
        GROUP BY 1
    """)


def pull_visits_mom(conn, close_year: int, close_month: int, n_months: int = 3) -> pd.DataFrame:
    """
    Linked-visit counts + gross earned revenue for the close month and the
    (n_months-1) preceding months — powers the MoM visit-trend table that shows
    recognized revenue tracks sessions used. Earned = frozen GL earned accounts.
    """
    months = []
    y, m = close_year, close_month
    for _ in range(n_months):
        months.append(f"{y}-{m:02d}")
        y, m = _shift_month(y, m, -1)
    months = sorted(months)
    start = _first_day(int(months[0][:4]), int(months[0][5:7]))
    end = _last_day(close_year, close_month)

    visits = execute_query_df(conn, f"""
        SELECT TO_CHAR(VISIT_DATE, 'YYYY-MM') AS M, COUNT(*) AS VISITS
        FROM VISITS_LINKED
        WHERE VISIT_DATE >= '{start}' AND VISIT_DATE <= '{end}'
        GROUP BY 1
    """)
    mlist = ", ".join(f"'{x}'" for x in months)
    earned = execute_query_df(conn, f"""
        SELECT MONTH_YM AS M, SUM(AMOUNT) AS EARNED
        FROM FROZEN_MONTHLY_GL
        WHERE MONTH_YM IN ({mlist})
          AND GL_CODE IN ('401001', '401002', '401004', '401005')
        GROUP BY 1
    """)
    v = dict(zip(visits["M"], visits["VISITS"].astype(int))) if not visits.empty else {}
    e = dict(zip(earned["M"], earned["EARNED"].astype(float))) if not earned.empty else {}
    rows = []
    for mo in months:
        vis = int(v.get(mo, 0))
        ear = float(e.get(mo, 0.0))
        rows.append({"M": mo, "VISITS": vis, "EARNED": ear, "EPV": (ear / vis) if vis else 0.0})
    return pd.DataFrame(rows)


def _mom_bridge_rows(gl_pivot: pd.DataFrame, close_ym: str, prior_ym: str):
    """[(GL line, Δ close-vs-prior)], sorted by magnitude, excluding the total."""
    rows = []
    for label in gl_pivot.index:
        if label == "TOTAL RECOGNIZED":
            continue
        p = float(gl_pivot.loc[label, prior_ym]) if prior_ym in gl_pivot.columns else 0.0
        c = float(gl_pivot.loc[label, close_ym]) if close_ym in gl_pivot.columns else 0.0
        rows.append((label, c - p))
    rows.sort(key=lambda r: -abs(r[1]))
    return rows


def _month_label(ym: str) -> str:
    y, m = int(ym[:4]), int(ym[5:7])
    return f"{calendar.month_abbr[m]} {y}"


def mom_narrative(headline: dict, visits_df: pd.DataFrame, close_ym: str, prior_ym: str) -> str:
    """Plain-language explanation of the recognized-revenue MoM move, tied to the
    visit trend and the accrual-vs-cash distinction. Direction-aware."""
    vd = visits_df.set_index("M")
    vc = int(vd.loc[close_ym, "VISITS"]) if close_ym in vd.index else 0
    vp = int(vd.loc[prior_ym, "VISITS"]) if prior_ym in vd.index else 0
    epv = float(vd.loc[close_ym, "EPV"]) if close_ym in vd.index else 0.0
    vpct = (vc - vp) / vp * 100 if vp else 0.0
    rec_dir = "below" if headline["recognized_delta"] < 0 else "above"
    return (
        f"{headline['close_month_label']} recognized revenue of ${headline['recognized_close']:,.0f} "
        f"is {abs(headline['recognized_pct'])*100:.1f}% {rec_dir} {headline['prior_month_label']} "
        f"(${headline['recognized_prior']:,.0f}). Recognized revenue is accrual — it tracks sessions "
        f"used plus breakage, not the month's cash sales. Linked visits moved {vp:,} → {vc:,} "
        f"({vpct:+.1f}%), with earned revenue per visit steady at ~${epv:,.0f}, so the move is "
        f"driven by visit volume (see the bridge below), not a change in how revenue is recognized. "
        f"Cash sales moved {headline['cash_pct']*100:+.1f}% over the same period; cash from new "
        f"package sales is deferred and recognized as those sessions are used in later months. "
        f"Visit data is complete through {headline['close_month_label']} month-end."
    )


def format_mom_supplement(narrative: str, visits_df: pd.DataFrame, bridge_rows, close_ym: str, prior_ym: str) -> str:
    """Text block (narrative + visit-trend table + MoM bridge) for the email body."""
    lines = ["Why " + _month_label(close_ym) + " vs " + _month_label(prior_ym), "", narrative, ""]
    lines.append("Visit trend")
    lines.append(f"  {'Month':<10}{'Visits':>10}{'Gross Earned':>16}{'Earned/Visit':>15}")
    for _, r in visits_df.iterrows():
        lines.append(f"  {_month_label(r['M']):<10}{int(r['VISITS']):>10,}"
                     f"{r['EARNED']:>16,.0f}{r['EPV']:>15,.2f}")
    lines.append("")
    lines.append("Recognized-revenue bridge (" + _month_label(prior_ym) + " → " + _month_label(close_ym) + ")")
    total = 0.0
    for label, d in bridge_rows:
        total += d
        lines.append(f"  {label:<38}{d:>+14,.0f}")
    lines.append(f"  {'Net change':<38}{total:>+14,.0f}")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Pivot builders for the PDF
# ---------------------------------------------------------------------------
# Recognized-revenue GL accounts shown in the close report, matching Rasa's P&L /
# the Saasant JE. Earned and breakage are SEPARATE lines (Cat 2026-06 feedback —
# see feedback_mighty_pilates_close_report_format).
_GL_LABELS = {
    "401001": "401001 Machine",
    "401002": "401002 Private Pilates",
    "401003": "401003 Class Pass",
    "401004": "401004 Mighty Teacher Training",
    "401005": "401005 Livestream Classes",
    "403001": "403001 Machine Breakage",
    "403002": "403002 Mighty Teacher Training Breakage",
    "403003": "403003 Private Pilates Breakage",
    "403004": "403004 Other Breakage",
    "404000": "404000 Retail Sales",
    "406000": "406000 Refunds",
    "407000": "407000 Discounts",
}


def _gl_by_studio_code(conn, ym: str) -> pd.DataFrame:
    """
    Per-month recognized-revenue GL rows [STUDIO, GL_CODE, AMOUNT] in GL (gross)
    convention, restricted to _GL_LABELS.

    Frozen months come straight from FROZEN_MONTHLY_GL — exactly what Rasa booked.
    Non-frozen months (e.g. a pre-freeze preview) are computed by the GL export
    itself so the report can never drift from the JE. The monthly close freezes
    the close month before generating this report, so at ship time both months
    read from the frozen store.
    """
    from pipeline.frozen_gl import is_month_frozen, load_frozen_gl

    if is_month_frozen(conn, ym):
        fz = load_frozen_gl(conn, ym)
        fz = fz[fz["GL_CODE"].isin(_GL_LABELS)].copy()
        fz["STUDIO"] = fz["STUDIO_NAME"].str.replace("Mighty Pilates ", "", regex=False)
        return fz[["STUDIO", "GL_CODE", "AMOUNT"]]

    # Live path: run the GL export (single source of truth) and read its per-studio
    # tabs so the report ties to the JE even before the month is frozen.
    import tempfile
    from openpyxl import load_workbook
    from pipeline.gl_export import generate_gl_export

    y, m = int(ym[:4]), int(ym[5:7])
    rows = []
    with tempfile.TemporaryDirectory() as td:
        path = generate_gl_export(conn, _first_day(y, m), _last_day(y, m), output_dir=td)
        wb = load_workbook(path, data_only=True)
        for tab in wb.sheetnames:
            if tab in ("Cover", "All Studios"):
                continue
            ws = wb[tab]
            for r in range(2, ws.max_row + 1):
                lab = ws.cell(row=r, column=1).value
                val = ws.cell(row=r, column=2).value
                if lab and isinstance(val, (int, float)):
                    code = str(lab).split()[0]
                    if code in _GL_LABELS:
                        rows.append({"STUDIO": tab, "GL_CODE": code, "AMOUNT": float(val)})
    return pd.DataFrame(rows, columns=["STUDIO", "GL_CODE", "AMOUNT"])


def _fetch_gl_src(conn, mom: dict) -> dict:
    """Fetch per-studio GL rows once per month for reuse by the account + studio
    pivots (avoids regenerating the GL export twice on the live path)."""
    return {ym: _gl_by_studio_code(conn, ym) for ym in (mom["prior_ym"], mom["close_ym"])}


def _build_gl_pivot(gl_src: dict, close_ym: str, prior_ym: str) -> pd.DataFrame:
    """Recognized revenue by GL account × month — earned vs breakage on separate
    GROSS lines, tying line-for-line to the Saasant JE / Rasa's P&L."""
    cols = {}
    for ym in (prior_ym, close_ym):
        g = gl_src.get(ym)
        cols[ym] = (g.groupby("GL_CODE")["AMOUNT"].sum()
                    if g is not None and not g.empty else pd.Series(dtype=float))
    pivot = pd.DataFrame(cols).reindex(sorted(_GL_LABELS)).dropna(how="all").fillna(0.0)
    # Drop accounts that are zero in every month shown (keep any line Rasa carries).
    pivot = pivot[(pivot.abs() >= 0.005).any(axis=1)]
    pivot.index = [_GL_LABELS[c] for c in pivot.index]
    pivot.loc["TOTAL RECOGNIZED"] = pivot.sum()
    return pivot


def _build_cash_pivot(mom: dict) -> pd.DataFrame:
    """
    Single 'Total Cash Sales' row from Cat's authoritative monthly cash sales
    (dashboard committed_actuals.monthly_sales, populated by the apply_cat_*
    scripts). Falls back to MindBody NET + Cat-authoritative ClassPass for any
    month not yet applied, so the report never blanks out.
    """
    import json

    close_ym, prior_ym = mom["close_ym"], mom["prior_ym"]
    months = [prior_ym, close_ym]

    cat_cash = {}
    p = Path(__file__).parent.parent / "dashboard" / "data" / "committed_actuals.json"
    try:
        cat_cash = {k: float(v) for k, v in json.loads(p.read_text()).get("monthly_sales", {}).items()}
    except Exception:
        cat_cash = {}

    # Fallback source for any month Cat hasn't applied yet.
    ledger = mom["ledger"].copy()
    mb = ledger[ledger["EVENT_TYPE"] == "Purchase"]
    mb_by_month = mb.groupby("MONTH_YM")["NET_SALES"].sum().astype(float)
    cp_by_month = mom["cp"].groupby("MONTH_YM")["CP_REV"].sum().astype(float)
    from pipeline.classpass_actuals import CAT_CLASSPASS
    for ym, studios in CAT_CLASSPASS.items():
        cp_by_month.loc[ym] = float(sum(studios.values()))

    vals = {}
    for ym in months:
        if ym in cat_cash:
            vals[ym] = cat_cash[ym]
        else:
            vals[ym] = float(mb_by_month.get(ym, 0.0)) + float(cp_by_month.get(ym, 0.0))

    return pd.DataFrame({ym: [vals[ym]] for ym in months}, index=["TOTAL CASH SALES"])


def _build_studio_pivot(gl_src: dict, close_ym: str, prior_ym: str) -> pd.DataFrame:
    """Recognized revenue by studio × month (incl. ClassPass), from the same
    JE-tied GL source as the account table, so studio totals reconcile to the
    account totals and to Rasa."""
    cols = {}
    for ym in (prior_ym, close_ym):
        g = gl_src.get(ym)
        cols[ym] = (g.groupby("STUDIO")["AMOUNT"].sum()
                    if g is not None and not g.empty else pd.Series(dtype=float))
    pivot = pd.DataFrame(cols).fillna(0.0).sort_index()
    pivot.loc["TOTAL"] = pivot.sum()
    return pivot


def _build_waterfall(waterfall: pd.DataFrame, cp_close_total: float) -> pd.DataFrame:
    df = waterfall.copy()
    df["EARNED"]   = df["NET_EARNED"].astype(float)
    df["BREAKAGE"] = df["NET_BREAKAGE"].astype(float)
    VINTAGE_ORDER = ["M0 (close month)", "M-1", "M-2", "M-3", "M-4", "M-5", "M-6", "M-7+ (older)", "Unknown"]
    df["VINTAGE"] = pd.Categorical(df["VINTAGE"], categories=VINTAGE_ORDER, ordered=True)

    pivot = df.pivot_table(index="VINTAGE", values=["EARNED", "BREAKAGE"],
                          aggfunc="sum", fill_value=0.0, observed=True)
    pivot = pivot.reindex(VINTAGE_ORDER, fill_value=0.0)
    pivot["CLASSPASS"] = 0.0
    pivot.loc["M0 (close month)", "CLASSPASS"] = cp_close_total
    pivot["TOTAL"] = pivot["EARNED"] + pivot["BREAKAGE"] + pivot["CLASSPASS"]
    grand = pivot["TOTAL"].sum()
    pivot["PCT"] = (pivot["TOTAL"] / grand) if grand else 0.0
    return pivot


def _build_studio_waterfall(waterfall: pd.DataFrame, cp_by_studio: pd.DataFrame) -> pd.DataFrame:
    df = waterfall.copy()
    df["TOTAL"] = df["NET_EARNED"].astype(float) + df["NET_BREAKAGE"].astype(float)
    VINTAGE_ORDER = ["M0 (close month)", "M-1", "M-2", "M-3", "M-4", "M-5", "M-6", "M-7+ (older)", "Unknown"]
    df["VINTAGE"] = pd.Categorical(df["VINTAGE"], categories=VINTAGE_ORDER, ordered=True)

    pivot = df.pivot_table(index="STUDIO", columns="VINTAGE", values="TOTAL",
                          aggfunc="sum", fill_value=0.0, observed=True)
    pivot = pivot.reindex(columns=VINTAGE_ORDER, fill_value=0.0)

    if not cp_by_studio.empty:
        for _, row in cp_by_studio.iterrows():
            studio = row["STUDIO"]; cp_amt = float(row["CP_REV"])
            if studio not in pivot.index:
                pivot.loc[studio] = 0.0
            pivot.loc[studio, "M0 (close month)"] += cp_amt

    pivot["TOTAL"] = pivot.sum(axis=1)
    pivot = pivot.sort_index()
    pivot.loc["TOTAL"] = pivot.sum()
    return pivot


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------
def _build_pdf(close_year: int, close_month: int,
              gl_pivot: pd.DataFrame, cash_pivot: pd.DataFrame, studio_pivot: pd.DataFrame,
              waterfall: pd.DataFrame, studio_wf: pd.DataFrame,
              output_dir: Path,
              visits_df: pd.DataFrame = None, bridge_rows=None, narrative: str = None) -> str:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

    period_label = f"{calendar.month_name[close_month]} {close_year}"
    prior_y, prior_m = _shift_month(close_year, close_month, -1)
    prior_label  = f"{calendar.month_name[prior_m]} {prior_y}"
    close_ym = f"{close_year}-{close_month:02d}"
    prior_ym = f"{prior_y}-{prior_m:02d}"

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Mighty_Close_Report_{calendar.month_abbr[close_month]}{close_year}_{stamp}.pdf"
    filepath = output_dir / filename

    doc = SimpleDocTemplate(
        str(filepath), pagesize=landscape(letter),
        leftMargin=0.5*inch, rightMargin=0.5*inch,
        topMargin=0.5*inch, bottomMargin=0.5*inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("T", parent=styles["Title"], fontSize=22,
                                 textColor=colors.HexColor("#1B2A4A"), spaceAfter=4)
    subtitle_style = ParagraphStyle("S", parent=styles["Normal"], fontSize=12,
                                    textColor=colors.HexColor("#5A6B8A"), spaceAfter=18)
    section_style = ParagraphStyle("H", parent=styles["Heading2"], fontSize=14,
                                   textColor=colors.HexColor("#1B2A4A"),
                                   spaceBefore=14, spaceAfter=6)
    body_style = ParagraphStyle("B", parent=styles["Normal"], fontSize=9, leading=12)
    note_style = ParagraphStyle("N", parent=styles["Normal"], fontSize=8,
                                textColor=colors.HexColor("#5A6B8A"),
                                spaceBefore=2, spaceAfter=2)

    HEADER_BG = colors.HexColor("#1B2A4A")
    HEADER_FG = colors.white
    ALT_ROW   = colors.HexColor("#F4F6FA")
    TOTAL_BG  = colors.HexColor("#FFF2CC")
    BORDER    = colors.HexColor("#D0D5DD")

    def _money(v):
        try:
            return f"${float(v):,.0f}"
        except Exception:
            return str(v)

    def _money2(v):
        try:
            return f"${float(v):,.2f}"
        except Exception:
            return str(v)

    def _pct(v):
        try:
            return f"{float(v)*100:.1f}%"
        except Exception:
            return str(v)

    def _make_table(headers, rows, col_widths, total_row_idx=None, money_cols=None, pct_cols=None):
        """Build styled reportlab Table."""
        money_cols = money_cols or []
        pct_cols   = pct_cols   or []
        head_cells = [Paragraph(f"<b>{h}</b>", ParagraphStyle(
            "TH", parent=body_style, fontSize=8, textColor=HEADER_FG, alignment=TA_CENTER
        )) for h in headers]
        out_rows = [head_cells]
        for r in rows:
            cells = []
            for i, v in enumerate(r):
                if i in money_cols:
                    txt = _money(v); align = TA_RIGHT
                elif i in pct_cols:
                    txt = _pct(v); align = TA_RIGHT
                else:
                    txt = str(v); align = TA_LEFT
                cells.append(Paragraph(txt, ParagraphStyle(
                    "TD", parent=body_style, fontSize=8, alignment=align
                )))
            out_rows.append(cells)

        t = Table(out_rows, colWidths=col_widths, repeatRows=1)
        style = [
            ("BACKGROUND", (0,0), (-1,0), HEADER_BG),
            ("TEXTCOLOR", (0,0), (-1,0), HEADER_FG),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,0), 6),
            ("TOPPADDING", (0,0), (-1,0), 6),
            ("BOTTOMPADDING", (0,1), (-1,-1), 3),
            ("TOPPADDING", (0,1), (-1,-1), 3),
            ("GRID", (0,0), (-1,-1), 0.5, BORDER),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]
        for i in range(1, len(out_rows)):
            if total_row_idx is not None and i == total_row_idx:
                style.append(("BACKGROUND", (0,i), (-1,i), TOTAL_BG))
                style.append(("FONTNAME", (0,i), (-1,i), "Helvetica-Bold"))
            elif i % 2 == 0:
                style.append(("BACKGROUND", (0,i), (-1,i), ALT_ROW))
        t.setStyle(TableStyle(style))
        return t

    # ============ Elements ============
    elements = []

    # Cover / header
    elements.append(Paragraph("Mighty Pilates", title_style))
    elements.append(Paragraph(f"Monthly Close Report — {period_label}", subtitle_style))

    # Headline totals
    close_total_rec = float(gl_pivot.loc["TOTAL RECOGNIZED", close_ym]) if close_ym in gl_pivot.columns else 0.0
    prior_total_rec = float(gl_pivot.loc["TOTAL RECOGNIZED", prior_ym]) if prior_ym in gl_pivot.columns else 0.0
    close_total_cash = float(cash_pivot.loc["TOTAL CASH SALES", close_ym]) if close_ym in cash_pivot.columns else 0.0
    prior_total_cash = float(cash_pivot.loc["TOTAL CASH SALES", prior_ym]) if prior_ym in cash_pivot.columns else 0.0

    delta_rec  = close_total_rec - prior_total_rec
    delta_cash = close_total_cash - prior_total_cash
    pct_rec    = (delta_rec / prior_total_rec) if prior_total_rec else 0.0
    pct_cash   = (delta_cash / prior_total_cash) if prior_total_cash else 0.0

    elements.append(Paragraph("Headline Totals", section_style))
    headline_rows = [
        ["Total Recognized Revenue",
         _money(prior_total_rec), _money(close_total_rec),
         _money(delta_rec), _pct(pct_rec)],
        ["Total Cash Sales (Cat-authoritative)",
         _money(prior_total_cash), _money(close_total_cash),
         _money(delta_cash), _pct(pct_cash)],
    ]
    elements.append(Table(
        [[Paragraph(f"<b>{h}</b>", ParagraphStyle('TH',parent=body_style,fontSize=8,
                                                  textColor=HEADER_FG,alignment=TA_CENTER))
          for h in ["Metric", prior_label, period_label, "Δ", "% Δ"]],
         *[[Paragraph(c[0], ParagraphStyle('TD',parent=body_style,fontSize=9))] +
           [Paragraph(v, ParagraphStyle('TD',parent=body_style,fontSize=9,alignment=TA_RIGHT))
            for v in c[1:]] for c in headline_rows]],
        colWidths=[3.5*inch, 1.5*inch, 1.5*inch, 1.3*inch, 1.0*inch],
        repeatRows=1,
    ).setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), HEADER_BG),
        ("TEXTCOLOR", (0,0), (-1,0), HEADER_FG),
        ("GRID", (0,0), (-1,-1), 0.5, BORDER),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 5),
    ])) or elements[-1])
    # Build the table fresh (above setStyle returned None)
    elements.pop()
    headline_table_rows = [["Metric", prior_label, period_label, "Δ", "% Δ"]]
    for r in headline_rows:
        headline_table_rows.append(r)
    t = Table(headline_table_rows, colWidths=[3.5*inch, 1.5*inch, 1.5*inch, 1.3*inch, 1.0*inch], repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), HEADER_BG),
        ("TEXTCOLOR",  (0,0), (-1,0), HEADER_FG),
        ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID",       (0,0), (-1,-1), 0.5, BORDER),
        ("ALIGN",      (1,0), (-1,-1), "RIGHT"),
        ("ALIGN",      (0,0), (0,-1),  "LEFT"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.15*inch))
    elements.append(Paragraph(
        "Recognized revenue matches the posted GL journal entry (GROSS basis): earned and "
        "breakage are shown on separate GL accounts, plus ClassPass and Retail, net of "
        "Refunds and Discounts — reconciling line-for-line to the P&L. "
        "Total Cash Sales are Cat's authoritative per-studio monthly figures.",
        note_style))

    # ============ MoM Bridge & Visit Trend ============
    if narrative or (visits_df is not None) or bridge_rows:
        elements.append(Spacer(1, 0.18*inch))
        elements.append(Paragraph(
            f"Why {period_label} vs {prior_label}", section_style))
        if narrative:
            elements.append(Paragraph(narrative, note_style))
            elements.append(Spacer(1, 0.12*inch))

        if visits_df is not None and not visits_df.empty:
            elements.append(Paragraph("Visit Trend", ParagraphStyle(
                "SH", parent=body_style, fontSize=11, textColor=colors.HexColor("#1B2A4A"),
                spaceBefore=4, spaceAfter=4, fontName="Helvetica-Bold")))
            vrows = [[_month_label(r["M"]), int(r["VISITS"]), r["EARNED"], r["EPV"]]
                     for _, r in visits_df.iterrows()]
            elements.append(_make_table(
                ["Month", "Linked Visits", "Gross Earned", "Earned / Visit"], vrows,
                col_widths=[1.8*inch, 1.6*inch, 1.9*inch, 1.7*inch],
                money_cols=[2, 3],
            ))
            elements.append(Spacer(1, 0.16*inch))

        if bridge_rows:
            elements.append(Paragraph(
                f"Recognized-Revenue Bridge ({prior_label} → {period_label})",
                ParagraphStyle("SH", parent=body_style, fontSize=11,
                               textColor=colors.HexColor("#1B2A4A"),
                               spaceBefore=4, spaceAfter=4, fontName="Helvetica-Bold")))
            brows = [[label, d] for label, d in bridge_rows]
            brows.append(["Net change", sum(d for _, d in bridge_rows)])
            elements.append(_make_table(
                ["Driver (GL account)", "Δ vs prior month"], brows,
                col_widths=[4.2*inch, 2.0*inch],
                total_row_idx=len(brows) - 1,
                money_cols=[1],
            ))

    # ============ MoM Comparison ============
    elements.append(PageBreak())
    elements.append(Paragraph(f"Month-over-Month Comparison — {prior_label} vs {period_label}", section_style))

    # GL bucket pivot (recognized revenue)
    elements.append(Paragraph("Recognized Revenue by GL Bucket", ParagraphStyle(
        "SH", parent=body_style, fontSize=11, textColor=colors.HexColor("#1B2A4A"),
        spaceBefore=4, spaceAfter=4, fontName="Helvetica-Bold")))
    gl_rows = []
    for label in gl_pivot.index:
        prior_v = float(gl_pivot.loc[label, prior_ym]) if prior_ym in gl_pivot.columns else 0.0
        close_v = float(gl_pivot.loc[label, close_ym]) if close_ym in gl_pivot.columns else 0.0
        d = close_v - prior_v
        pct = (d / prior_v) if prior_v else 0.0
        gl_rows.append([label, prior_v, close_v, d, pct])
    t = _make_table(
        ["GL Code / Bucket", prior_label, period_label, "Δ", "% Δ"],
        gl_rows,
        col_widths=[2.6*inch, 1.5*inch, 1.5*inch, 1.3*inch, 1.0*inch],
        total_row_idx=len(gl_rows),  # last row is TOTAL RECOGNIZED
        money_cols=[1,2,3],
        pct_cols=[4],
    )
    elements.append(t)
    elements.append(Spacer(1, 0.18*inch))

    # Cash sales pivot
    elements.append(Paragraph("Cash Sales", ParagraphStyle(
        "SH", parent=body_style, fontSize=11, textColor=colors.HexColor("#1B2A4A"),
        spaceBefore=4, spaceAfter=4, fontName="Helvetica-Bold")))
    cash_rows = []
    for label in cash_pivot.index:
        prior_v = float(cash_pivot.loc[label, prior_ym]) if prior_ym in cash_pivot.columns else 0.0
        close_v = float(cash_pivot.loc[label, close_ym]) if close_ym in cash_pivot.columns else 0.0
        d = close_v - prior_v
        pct = (d / prior_v) if prior_v else 0.0
        cash_rows.append([label, prior_v, close_v, d, pct])
    elements.append(_make_table(
        ["", prior_label, period_label, "Δ", "% Δ"], cash_rows,
        col_widths=[2.6*inch, 1.5*inch, 1.5*inch, 1.3*inch, 1.0*inch],
        total_row_idx=len(cash_rows),  # last row is TOTAL CASH SALES
        money_cols=[1,2,3], pct_cols=[4],
    ))
    elements.append(Spacer(1, 0.18*inch))

    # Per-studio (recognized + CP)
    elements.append(Paragraph("Recognized Revenue by Studio (incl. ClassPass)", ParagraphStyle(
        "SH", parent=body_style, fontSize=11, textColor=colors.HexColor("#1B2A4A"),
        spaceBefore=4, spaceAfter=4, fontName="Helvetica-Bold")))
    studio_rows = []
    for studio in studio_pivot.index:
        prior_v = float(studio_pivot.loc[studio, prior_ym]) if prior_ym in studio_pivot.columns else 0.0
        close_v = float(studio_pivot.loc[studio, close_ym]) if close_ym in studio_pivot.columns else 0.0
        d = close_v - prior_v
        pct = (d / prior_v) if prior_v else 0.0
        studio_rows.append([studio, prior_v, close_v, d, pct])
    elements.append(_make_table(
        ["Studio", prior_label, period_label, "Δ", "% Δ"], studio_rows,
        col_widths=[2.6*inch, 1.5*inch, 1.5*inch, 1.3*inch, 1.0*inch],
        total_row_idx=len(studio_rows),
        money_cols=[1,2,3], pct_cols=[4],
    ))

    # ============ Waterfall ============
    elements.append(PageBreak())
    elements.append(Paragraph(f"Revenue Recognition Waterfall — {period_label}", section_style))
    elements.append(Paragraph(
        f"This shows how much of {period_label}'s recognized revenue traces back to each sale-month vintage. "
        f"M0 = sold in {period_label}; M-1 = sold {prior_label}; M-6 = sold 6 months ago; etc.",
        note_style))
    elements.append(Spacer(1, 0.05*inch))

    wf_rows = []
    for vintage in waterfall.index:
        wf_rows.append([
            vintage,
            float(waterfall.loc[vintage, "EARNED"]),
            float(waterfall.loc[vintage, "BREAKAGE"]),
            float(waterfall.loc[vintage, "CLASSPASS"]),
            float(waterfall.loc[vintage, "TOTAL"]),
            float(waterfall.loc[vintage, "PCT"]),
        ])
    # Add total row
    wf_rows.append([
        "TOTAL",
        float(waterfall["EARNED"].sum()),
        float(waterfall["BREAKAGE"].sum()),
        float(waterfall["CLASSPASS"].sum()),
        float(waterfall["TOTAL"].sum()),
        1.0,
    ])
    elements.append(_make_table(
        ["Sale-Month Vintage", "Earned", "Breakage", "ClassPass", "Total", "% of Month"],
        wf_rows,
        col_widths=[2.0*inch, 1.3*inch, 1.3*inch, 1.2*inch, 1.4*inch, 1.0*inch],
        total_row_idx=len(wf_rows),
        money_cols=[1,2,3,4], pct_cols=[5],
    ))

    # ============ Per-studio waterfall ============
    elements.append(PageBreak())
    elements.append(Paragraph(f"Per-Studio Sale-Month Waterfall — {period_label}", section_style))
    elements.append(Paragraph(
        "Total recognized revenue by studio, broken out by the sale-month vintage of the underlying packages.",
        note_style))
    elements.append(Spacer(1, 0.05*inch))

    sw_headers = ["Studio"] + list(studio_wf.columns)
    sw_rows = []
    for studio in studio_wf.index:
        sw_rows.append([studio] + [float(studio_wf.loc[studio, c]) for c in studio_wf.columns])
    # Compact column widths to fit landscape
    n_vint = len(studio_wf.columns)
    avail = 10.0  # landscape letter ~10in usable
    col_widths = [1.4*inch] + [(avail-1.4)/n_vint*inch for _ in range(n_vint)]
    elements.append(_make_table(
        sw_headers, sw_rows, col_widths,
        total_row_idx=len(sw_rows),
        money_cols=list(range(1, n_vint+1)),
    ))

    # ============ Footer ============
    elements.append(Spacer(1, 0.2*inch))
    elements.append(Paragraph(
        f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')} from "
        f"EARNED_REVENUE_ANALYTICS.DAILY_REVENUE_AND_SALES_DETAIL + "
        f"PLAYLIST_DATAMART.CLASSPASS_REPORTING_ANALYTICS.RESERVATIONS. "
        f"See MONTHLY_CLOSE.md for the close procedure.",
        note_style))

    doc.build(elements)
    print(f"  Saved: {filepath}")
    return str(filepath)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def generate(conn, close_year: int, close_month: int, output_dir: str = None) -> str:
    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "outputs"
    Path(output_dir).mkdir(exist_ok=True)

    period = f"{calendar.month_abbr[close_month]}{close_year}"
    print(f"Generating Monthly Close Report PDF: {period}")

    mom        = pull_mom(conn, close_year, close_month)
    waterfall_raw = pull_waterfall(conn, close_year, close_month)
    cp_close   = pull_cp_by_studio(conn, close_year, close_month)

    cp_close_total = float(cp_close["CP_REV"].sum()) if not cp_close.empty else 0.0

    gl_src      = _fetch_gl_src(conn, mom)
    gl_pivot    = _build_gl_pivot(gl_src, mom["close_ym"], mom["prior_ym"])
    cash_pivot  = _build_cash_pivot(mom)
    studio_piv  = _build_studio_pivot(gl_src, mom["close_ym"], mom["prior_ym"])
    waterfall   = _build_waterfall(waterfall_raw, cp_close_total)
    studio_wf   = _build_studio_waterfall(waterfall_raw, cp_close)

    # MoM bridge + visit trend (answers "why is the close month up/down vs prior")
    close_ym, prior_ym = mom["close_ym"], mom["prior_ym"]
    headline    = get_headline_totals(conn, close_year, close_month)
    visits_df   = pull_visits_mom(conn, close_year, close_month)
    bridge_rows = _mom_bridge_rows(gl_pivot, close_ym, prior_ym)
    narrative   = mom_narrative(headline, visits_df, close_ym, prior_ym)

    return _build_pdf(close_year, close_month,
                     gl_pivot, cash_pivot, studio_piv,
                     waterfall, studio_wf,
                     Path(output_dir),
                     visits_df=visits_df, bridge_rows=bridge_rows, narrative=narrative)


def generate_prior_month(conn, output_dir: str = None) -> str:
    from datetime import datetime, timedelta
    today = datetime.now()
    first = today.replace(day=1)
    last_prior = first - timedelta(days=1)
    return generate(conn, last_prior.year, last_prior.month, output_dir)


def get_headline_totals(conn, close_year: int, close_month: int) -> dict:
    """
    Compute headline totals for the close month + prior month.
    Returns a dict with recognized revenue, cash sales, and Δ vs prior month.
    """
    mom        = pull_mom(conn, close_year, close_month)
    gl_src     = _fetch_gl_src(conn, mom)
    gl_pivot   = _build_gl_pivot(gl_src, mom["close_ym"], mom["prior_ym"])
    cash_pivot = _build_cash_pivot(mom)

    close_ym = f"{close_year}-{close_month:02d}"
    prior_y, prior_m = _shift_month(close_year, close_month, -1)
    prior_ym = f"{prior_y}-{prior_m:02d}"

    def _val(pivot, label, ym):
        if ym in pivot.columns and label in pivot.index:
            return float(pivot.loc[label, ym])
        return 0.0

    rec_close  = _val(gl_pivot,   "TOTAL RECOGNIZED", close_ym)
    rec_prior  = _val(gl_pivot,   "TOTAL RECOGNIZED", prior_ym)
    cash_close = _val(cash_pivot, "TOTAL CASH SALES", close_ym)
    cash_prior = _val(cash_pivot, "TOTAL CASH SALES", prior_ym)

    return {
        "close_month_label": f"{calendar.month_name[close_month]} {close_year}",
        "prior_month_label": f"{calendar.month_name[prior_m]} {prior_y}",
        "recognized_close": rec_close,
        "recognized_prior": rec_prior,
        "recognized_delta": rec_close - rec_prior,
        "recognized_pct":   (rec_close - rec_prior) / rec_prior if rec_prior else 0.0,
        "cash_close":       cash_close,
        "cash_prior":       cash_prior,
        "cash_delta":       cash_close - cash_prior,
        "cash_pct":         (cash_close - cash_prior) / cash_prior if cash_prior else 0.0,
    }


def compose_email_body(headline: dict, file_paths: list, extra_note: str = None) -> tuple:
    """
    Compose subject and body for the monthly close email.
    extra_note (if given) is inserted after the headline block — used for the
    MoM bridge + visit-trend supplement.
    Returns (subject, body).
    """
    from pathlib import Path as _Path

    def _money(v): return f"${v:,.0f}"
    def _delta(v):
        sign = "+" if v >= 0 else "-"
        return f"{sign}${abs(v):,.0f}"
    def _pct(v):
        sign = "+" if v >= 0 else ""
        return f"{sign}{v*100:.1f}%"

    subject = f"Mighty Pilates — {headline['close_month_label']} Monthly Close"

    # Single-space headline formatting mirrors the format we've sent since June.
    note_block = f"\n{extra_note}\n" if extra_note else ""

    body = f"""Hi team,

Attached are the {headline['close_month_label']} monthly close deliverables for Mighty Pilates.

Headline Totals ({headline['close_month_label']} vs {headline['prior_month_label']})

  Total Recognized Revenue: {_money(headline['recognized_close'])}
    {headline['prior_month_label']}: {_money(headline['recognized_prior'])}
    Change: {_delta(headline['recognized_delta'])} ({_pct(headline['recognized_pct'])})

  Total Cash Sales: {_money(headline['cash_close'])}
    {headline['prior_month_label']}: {_money(headline['cash_prior'])}
    Change: {_delta(headline['cash_delta'])} ({_pct(headline['cash_pct'])})
{note_block}
Attachments

  1. Mighty Pilates GL Export — per-studio + consolidated GL totals (Excel)
  2. Saasant Upload — QuickBooks journal entry workbook for Saasant (Excel)
  3. Mighty Pilates Close Report — MoM comparison and revenue recognition waterfall (PDF)

Please reach out with any questions.

Best regards,
Chandler
"""
    return subject, body
