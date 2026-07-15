"""Helpers for parsing QuickBooks Online Excel exports.

This module is the union of the patterns observed in CNS and Alma Mater:

* Low-level header / row helpers that don't care about chart of accounts —
  ``parse_qbo_headers`` (fixed-row, fast), ``detect_header_row`` (autodetect,
  resilient), ``parse_month_header`` (regex + datetime + "Mon-YY"),
  ``find_row_by_search``, ``extract_row_values``, ``find_pl_sheet``,
  ``find_bs_sheet``.
* Account-mapping primitives for clients that iterate every account row and
  classify against a registry — ``normalize_key``, ``extract_account_code``,
  ``map_account``.

Each client's chart-of-accounts registry (``CNS QBO_ACCOUNTS``, ``Alma
QBO_PL_SEARCH``) stays in the client's own repo. This module only provides
the resolution logic.
"""

from __future__ import annotations

import re
from datetime import datetime
from typing import Any, Iterable, Optional

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from empirica_core.months import MONTH_MAP

# ---------------------------------------------------------------------------
# Regex tokens
# ---------------------------------------------------------------------------

MONTH_TOKEN_RE = re.compile(
    r"\b("
    r"jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|"
    r"nov(?:ember)?|dec(?:ember)?"
    r")\b",
    re.IGNORECASE,
)
YEAR_RE = re.compile(r"\b(20\d{2})\b")
ACCOUNT_CODE_RE = re.compile(r"^\s*(\d{3,4})\b")


# ---------------------------------------------------------------------------
# Month header parsing
# ---------------------------------------------------------------------------

def parse_month_header(cell_value: Any) -> Optional[tuple[int, int]]:
    """Parse a header cell into ``(year, month_num)`` or ``None``.

    Accepts: ``datetime`` objects, ``'January 2025'``, ``'Jan 2026'``,
    ``'Jan-26'``, ``'Jan-2026'``. Returns ``None`` for blanks, ``'Total'``,
    or anything that doesn't contain both a month token and a 4-digit year
    (or 2-digit year in the ``Mon-YY`` shape).
    """
    if cell_value is None:
        return None
    if isinstance(cell_value, datetime):
        return (cell_value.year, cell_value.month)
    s = str(cell_value).strip()
    if not s or s.lower() == "total":
        return None

    m = MONTH_TOKEN_RE.search(s)
    if not m:
        return None
    month_str = m.group(1).lower()
    # Try the full token first ('september' → 9), then the 3-letter prefix
    # ('sept' / 'sep' both fall back to 'sep').
    if month_str in MONTH_MAP:
        month_num = MONTH_MAP[month_str]
    elif month_str[:3] in MONTH_MAP:
        month_num = MONTH_MAP[month_str[:3]]
    else:
        return None

    y = YEAR_RE.search(s)
    if y:
        return (int(y.group(1)), month_num)

    # Fall back to 2-digit year (e.g. 'Jan-26' → assume 20xx)
    yy = re.search(r"-(\d{2})\b", s)
    if yy:
        return (2000 + int(yy.group(1)), month_num)
    return None


def parse_qbo_headers(
    ws: Worksheet,
    header_row: int = 5,
    max_col: int = 50,
) -> dict[tuple[int, int], int]:
    """Return ``{(year, month_num): column_index}`` from a fixed header row.

    Fast path for QBO exports whose layout you know — defaults to row 5,
    QBO's standard for P&L / BS sheets. Stops at the ``Total`` column.
    """
    headers: dict[tuple[int, int], int] = {}
    for c in range(2, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            continue
        s = str(val).strip()
        if s == "" or s.lower() == "total":
            break
        parsed = parse_month_header(val)
        if parsed is not None:
            headers[parsed] = c
    return headers


def detect_header_row(
    ws: Worksheet,
    max_scan: int = 12,
    max_col: int = 50,
) -> tuple[int, dict[tuple[int, int], int]]:
    """Scan rows ``1..max_scan`` for the row containing month-shaped headers.

    Returns ``(row_index, headers)``. Use this when the header row position
    varies across files. Raises ``ValueError`` if no header row is found.
    """
    last_row = min(ws.max_row or max_scan, max_scan)
    for r in range(1, last_row + 1):
        candidates: dict[tuple[int, int], int] = {}
        for c in range(2, max_col + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            if str(val).strip().lower() == "total":
                break
            parsed = parse_month_header(val)
            if parsed is not None:
                candidates[parsed] = c
        if candidates:
            return r, candidates
    raise ValueError("Could not locate a month header row in the worksheet.")


# ---------------------------------------------------------------------------
# Row helpers
# ---------------------------------------------------------------------------

def find_row_by_search(
    ws: Worksheet,
    search_terms: Iterable[str],
    max_row: int = 200,
    column: int = 1,
) -> Optional[int]:
    """Find the first row whose ``column`` cell contains any ``search_terms``.

    Matching is case-insensitive substring. Returns ``None`` if nothing
    matches.
    """
    terms = [t.lower() for t in search_terms]
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=column).value
        if val is None:
            continue
        val_lower = str(val).strip().lower()
        for t in terms:
            if t in val_lower:
                return r
    return None


def extract_row_values(
    ws: Worksheet,
    row: int,
    headers: dict[tuple[int, int], int],
) -> dict[tuple[int, int], float]:
    """Pull numeric values from ``row``, keyed by the ``(year, month)`` map.

    ``None`` and non-numeric cells become ``0.0`` so downstream math doesn't
    have to special-case blanks.
    """
    out: dict[tuple[int, int], float] = {}
    for (yr, mo), col in headers.items():
        val = ws.cell(row=row, column=col).value
        try:
            out[(yr, mo)] = float(val) if val is not None else 0.0
        except (TypeError, ValueError):
            out[(yr, mo)] = 0.0
    return out


# ---------------------------------------------------------------------------
# Sheet detection
# ---------------------------------------------------------------------------

def find_pl_sheet(wb: Workbook) -> Optional[str]:
    """Locate the Profit & Loss sheet in a QBO export workbook."""
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if nl in ("pl", "p&l", "p & l"):
            return name
        if "profit" in nl and "loss" in nl:
            return name
    return None


def find_bs_sheet(wb: Workbook) -> Optional[str]:
    """Locate the Balance Sheet sheet in a QBO export workbook."""
    for name in wb.sheetnames:
        nl = name.lower().strip()
        if nl == "bs":
            return name
        if "balance" in nl and "sheet" in nl:
            return name
    return None


# ---------------------------------------------------------------------------
# Account mapping
# ---------------------------------------------------------------------------

def extract_account_code(label: Any) -> Optional[str]:
    """Pull the leading 3-4 digit account code off a label, if present.

    ``'400 Fee Income'`` → ``'400'``; ``'Net Income'`` → ``None``.
    """
    if label is None:
        return None
    m = ACCOUNT_CODE_RE.match(str(label))
    return m.group(1) if m else None


def normalize_key(label: Any) -> str:
    """Convert an account label into a snake_case key.

    Strips a leading 3-4 digit code, removes ``@`` compensation tags and
    parentheticals, lowercases, collapses runs of non-alphanumerics to ``_``.
    """
    if label is None:
        return "unmapped"
    s = str(label).strip().lower()
    s = re.sub(r"^\d{3,4}\s+", "", s)
    s = re.split(r"\s+@\s+", s)[0]
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    s = s.strip("_")
    return s or "unmapped"


def map_account(
    label: Any,
    registry: dict[str, dict],
    extras: Optional[dict[str, str]] = None,
) -> dict:
    """Resolve an account label against a chart-of-accounts ``registry``.

    ``registry`` shape: ``{code: {"key": str, "name": str, ...}}`` — same as
    CNS's ``QBO_ACCOUNTS``. ``extras`` is an optional ``{raw_label: target_key}``
    override map for one-off mappings the user has added through a UI.

    Resolution order:
        1. ``extras`` exact-label match
        2. ``registry`` lookup by 3-4 digit code prefix
        3. ``registry`` exact name match (code-stripped, lowercased)
        4. ``registry`` fuzzy contains match
        5. Unmapped — returns a ``normalize_key``-derived suggestion

    Returns a dict: ``{code, key, name, mapped, source}`` where ``source`` is
    one of ``"extras" | "code" | "name" | "fuzzy" | "suggested" | "empty"``.
    """
    if label is None or str(label).strip() == "":
        return {"code": None, "key": None, "name": "", "mapped": False, "source": "empty"}

    raw = str(label).strip()
    extras = extras or {}

    if raw in extras:
        return {
            "code": extract_account_code(raw),
            "key": extras[raw],
            "name": raw,
            "mapped": True,
            "source": "extras",
        }

    code = extract_account_code(raw)
    if code and code in registry:
        meta = registry[code]
        return {
            "code": code,
            "key": meta["key"],
            "name": meta.get("name", raw),
            "mapped": True,
            "source": "code",
        }

    name_only = re.sub(r"^\s*\d{3,4}\s+", "", raw).lower().strip()
    by_name = {meta["name"].strip().lower(): (c, meta) for c, meta in registry.items()
               if "name" in meta}

    if name_only in by_name:
        c, meta = by_name[name_only]
        return {"code": c, "key": meta["key"], "name": meta["name"],
                "mapped": True, "source": "name"}

    for name_lc, (c, meta) in by_name.items():
        if name_lc and name_lc in name_only:
            return {"code": c, "key": meta["key"], "name": meta["name"],
                    "mapped": True, "source": "fuzzy"}

    return {
        "code": code,
        "key": normalize_key(raw),
        "name": raw,
        "mapped": False,
        "source": "suggested",
    }


__all__ = [
    "MONTH_TOKEN_RE",
    "YEAR_RE",
    "ACCOUNT_CODE_RE",
    "parse_month_header",
    "parse_qbo_headers",
    "detect_header_row",
    "find_row_by_search",
    "extract_row_values",
    "find_pl_sheet",
    "find_bs_sheet",
    "extract_account_code",
    "normalize_key",
    "map_account",
]
